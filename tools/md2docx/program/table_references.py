from __future__ import annotations

import re
import shlex
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


_REFERENCE_PATTERN = re.compile(r"\{\{\s*table-ref\b(.*?)\}\}", re.DOTALL)
_REFERENCE_START = re.compile(r"\{\{\s*table-ref\b")


@dataclass(frozen=True, slots=True)
class TableReference:
    marker: str
    source: Path
    sheet: str | None
    cell_range: str | None
    table_id: str | None
    ordinal: int


@dataclass(frozen=True, slots=True)
class ExternalTableCell:
    value: str
    fill: str | None
    font_color: str | None
    bold: bool
    italic: bool
    font_size: float | None
    horizontal: str | None
    vertical: str | None
    borders: dict[str, tuple[str, str, int]]


@dataclass(frozen=True, slots=True)
class ExternalTable:
    reference: TableReference
    rows: list[list[ExternalTableCell]]
    merges: list[tuple[int, int, int, int]]
    column_widths: list[float]
    row_heights: list[float | None]


def extract_references(markdown: str, *, base_path: Path) -> tuple[str, list[TableReference]]:
    """Replace reference blocks with stable line markers before Markdown parsing."""
    references: list[TableReference] = []

    def replace(match: re.Match[str]) -> str:
        values = _parse_attributes(match.group(1))
        source_value = values.get("source", "")
        if not source_value:
            raise ValueError("table-ref 缺少 source 属性。")
        source = Path(source_value).expanduser()
        if not source.is_absolute():
            source = base_path / source
        source = source.resolve(strict=False)
        reference = TableReference(
            marker=f"__TIANCE_TABLE_REF_{len(references)}__",
            source=source,
            sheet=values.get("sheet") or None,
            cell_range=values.get("range") or None,
            table_id=values.get("table_id") or None,
            ordinal=len(references) + 1,
        )
        references.append(reference)
        return f"\n{reference.marker}\n"

    content = _REFERENCE_PATTERN.sub(replace, markdown)
    if _REFERENCE_START.search(content):
        raise ValueError("table-ref 指令未闭合，必须以 }} 结束。")
    return content, references


def load_external_table(reference: TableReference) -> ExternalTable:
    if not reference.source.is_file():
        raise ValueError(f"table-ref 来源文件不存在：{reference.source}")
    if reference.source.suffix.lower() != ".xlsx":
        raise ValueError("当前 table-ref 只支持 .xlsx 文件。")
    workbook = load_workbook(reference.source, data_only=False, read_only=False)
    try:
        sheet_name = reference.sheet or workbook.sheetnames[0]
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Excel 工作表不存在：{sheet_name}")
        worksheet = workbook[sheet_name]
        min_col, min_row, max_col, max_row = _resolve_bounds(worksheet, reference.cell_range)
        rows: list[list[ExternalTableCell]] = []
        for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            rows.append([_cell_from_excel(cell) for cell in row])

        merges = []
        for merged in worksheet.merged_cells.ranges:
            left, top, right, bottom = range_boundaries(str(merged))
            if left >= min_col and top >= min_row and right <= max_col and bottom <= max_row:
                merges.append((top - min_row, left - min_col, bottom - min_row, right - min_col))

        widths = [
            max(24.0, float(worksheet.column_dimensions[get_column_letter(column)].width or 12.0) * 5.2)
            for column in range(min_col, max_col + 1)
        ]
        heights = [worksheet.row_dimensions[row].height for row in range(min_row, max_row + 1)]
        return ExternalTable(reference, rows, merges, widths, heights)
    finally:
        workbook.close()


def _parse_attributes(raw: str) -> dict[str, str]:
    try:
        tokens = shlex.split(raw.replace("\n", " "), posix=True)
    except ValueError as exc:
        raise ValueError(f"table-ref 属性格式错误：{exc}") from exc
    values: dict[str, str] = {}
    for token in tokens:
        if "=" not in token:
            raise ValueError(f"table-ref 属性必须使用 key=\"value\"：{token}")
        key, value = token.split("=", 1)
        if key not in {"source", "sheet", "range", "table_id"} or not value:
            raise ValueError(f"table-ref 不支持或缺少值的属性：{key}")
        values[key] = value
    return values


def _resolve_bounds(worksheet: Any, cell_range: str | None) -> tuple[int, int, int, int]:
    if cell_range:
        try:
            return range_boundaries(cell_range)
        except ValueError as exc:
            raise ValueError(f"Excel range 格式错误：{cell_range}") from exc
    dimension = worksheet.calculate_dimension()
    return range_boundaries(dimension)


def _cell_from_excel(cell: Any) -> ExternalTableCell:
    value = "" if cell.value is None else str(cell.value)
    fill = _color(cell.fill.fgColor) if cell.fill.fill_type else None
    font_color = _color(cell.font.color)
    return ExternalTableCell(
        value=value,
        fill=fill,
        font_color=font_color,
        bold=bool(cell.font.bold),
        italic=bool(cell.font.italic),
        font_size=float(cell.font.sz) if cell.font.sz else None,
        horizontal=cell.alignment.horizontal,
        vertical=cell.alignment.vertical,
        borders=_borders_from_excel(cell),
    )


def _color(color: Any) -> str | None:
    if color is None:
        return None
    value = getattr(color, "rgb", None)
    if isinstance(value, str):
        value = value[-6:]
        if re.fullmatch(r"[0-9A-Fa-f]{6}", value):
            return value.upper()
    return None


def _borders_from_excel(cell: Any) -> dict[str, tuple[str, str, int]]:
    result: dict[str, tuple[str, str, int]] = {}
    for name in ("top", "left", "bottom", "right"):
        side = getattr(cell.border, name)
        if not side or not side.style:
            continue
        result[name] = (
            _word_border_style(side.style),
            _color(side.color) or "000000",
            {"thin": 4, "medium": 8, "thick": 12}.get(side.style, 4),
        )
    return result


def _word_border_style(value: str) -> str:
    return {
        "thin": "single",
        "medium": "single",
        "thick": "single",
        "double": "double",
        "dashed": "dashed",
        "dotted": "dotted",
        "dashDot": "dashDot",
        "dashDotDot": "dashDotDot",
        "hair": "single",
    }.get(value, "single")
