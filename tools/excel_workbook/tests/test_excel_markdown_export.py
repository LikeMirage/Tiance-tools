from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


PROGRAM = Path(__file__).parents[1] / "program"
sys.path.insert(0, str(PROGRAM))

from excel_markdown_export import export_workbook  # noqa: E402


def test_export_writes_content_format_and_rules(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    content = tmp_path / "content.md"
    format_file = tmp_path / "format.md"
    report = tmp_path / "report.md"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销售"
    sheet.append(["区域", "销售额", "完成率"])
    sheet.append(["华东", 120, 0.8])
    sheet.append(["华南", 98, 0.7])
    sheet.merge_cells("A1:C1")
    sheet["A1"].font = Font(bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="4472C4")
    sheet["B2"].number_format = "#,##0"
    sheet["C2"].number_format = "0.0%"
    workbook.save(source)

    result = export_workbook(source, content, format_file, tmp_path / "assets", report)

    assert result["protocol"] == "md2xlsx-dual-markdown"
    assert result["rules"]["round_trip"]
    assert "华东" in content.read_text(encoding="utf-8")
    format_text = format_file.read_text(encoding="utf-8")
    assert "merge:merge1" in format_text
    assert "font-weight:bold" in format_text
    report_text = report.read_text(encoding="utf-8")
    assert "先读取 content.md" in report_text


def test_exported_markdown_can_be_read_back(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["名称", "值"])
    sheet.append(["A", 1])
    workbook.save(source)

    result = export_workbook(source, tmp_path / "content.md", tmp_path / "format.md", tmp_path / "assets", None)
    assert result["sheet_count"] == 1
    assert load_workbook(source).sheetnames == ["数据"]
