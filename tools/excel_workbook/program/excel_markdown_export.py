from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter, range_boundaries


RULES = {
    "content": "content.md 只保存每个 Sheet 的正文表格、数字、日期、布尔值和 Excel 公式。",
    "format": "format.md 与 content.md 同标题、同行数、同列数；每个格式单元格使用分号分隔的格式指令。",
    "merge": "合并区域的所有格使用相同 merge:名称，正文只保留左上角内容。",
    "round_trip": "再次转换的目标是结构和格式语义等价，不承诺 Excel 文件字节级相同。",
}


def export_workbook(
    input_path: Path,
    content_path: Path,
    format_path: Path,
    assets_dir: Path,
    report_path: Path | None,
    include_empty_rows: bool = False,
    selected_sheets: list[str] | None = None,
) -> dict[str, Any]:
    workbook = load_workbook(input_path, data_only=False, read_only=False)
    warnings: list[str] = []
    assets: list[str] = []
    try:
        sheets = [sheet for sheet in workbook.worksheets if not selected_sheets or sheet.title in selected_sheets]
        if selected_sheets:
            missing = [name for name in selected_sheets if name not in workbook.sheetnames]
            if missing:
                raise ValueError(f"指定的 Sheet 不存在：{', '.join(missing)}")
        content_parts: list[str] = []
        format_parts: list[str] = []
        for sheet in sheets:
            bounds = _sheet_bounds(sheet, include_empty_rows)
            if bounds is None:
                warnings.append(f"{sheet.title}: 没有可导出的单元格，已跳过。")
                continue
            min_row, min_col, max_row, max_col = bounds
            content_parts.append(_render_content_table(sheet, min_row, min_col, max_row, max_col))
            image_map = _export_images(sheet, assets_dir, warnings)
            format_parts.append(_render_format_table(sheet, min_row, min_col, max_row, max_col, warnings, image_map, assets_dir))
            assets.extend(image_map.values())
            _collect_unsupported_objects(sheet, warnings)
        content_path.write_text("\n\n".join(content_parts).rstrip() + "\n", encoding="utf-8")
        format_path.write_text("\n\n".join(format_parts).rstrip() + "\n", encoding="utf-8")
        report = _render_report(input_path, content_path, format_path, sheets, warnings, assets)
        if report_path is not None:
            report_path.write_text(report, encoding="utf-8")
        return {
            "input_path": str(input_path),
            "content_path": str(content_path),
            "format_path": str(format_path),
            "report_path": str(report_path) if report_path else None,
            "assets_dir": str(assets_dir) if assets else None,
            "sheet_count": len(sheets),
            "sheets": [{"name": sheet.title, "range": _sheet_bounds_text(sheet, include_empty_rows)} for sheet in sheets],
            "asset_count": len(assets),
            "warnings": warnings,
            "rules": RULES,
            "protocol": "md2xlsx-dual-markdown",
        }
    finally:
        workbook.close()


def _sheet_bounds(sheet: Any, include_empty_rows: bool) -> tuple[int, int, int, int] | None:
    dimension = sheet.calculate_dimension()
    if not dimension or dimension == "A1:A1" and sheet["A1"].value is None and not sheet.merged_cells.ranges:
        return None
    min_col, min_row, max_col, max_row = range_boundaries(dimension)
    for image in getattr(sheet, "_images", []) or []:
        anchor = getattr(getattr(image, "anchor", None), "_from", None)
        if anchor is not None:
            max_row = max(max_row, anchor.row + 1)
            max_col = max(max_col, anchor.col + 1)
    if not include_empty_rows:
        while max_row > min_row and all(sheet.cell(max_row, col).value is None for col in range(min_col, max_col + 1)) and not _has_image_on_row(sheet, max_row):
            max_row -= 1
    return min_row, min_col, max_row, max_col


def _has_image_on_row(sheet: Any, row: int) -> bool:
    return any(
        getattr(getattr(image, "anchor", None), "_from", None) is not None
        and image.anchor._from.row + 1 == row
        for image in getattr(sheet, "_images", []) or []
    )


def _sheet_bounds_text(sheet: Any, include_empty_rows: bool) -> str | None:
    bounds = _sheet_bounds(sheet, include_empty_rows)
    if not bounds:
        return None
    min_row, min_col, max_row, max_col = bounds
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _render_content_table(sheet: Any, min_row: int, min_col: int, max_row: int, max_col: int) -> str:
    lines = [f"# {sheet.title}", ""]
    for row in range(min_row, max_row + 1):
        values = []
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row, col)
            if _merged_non_owner(sheet, cell.coordinate):
                value = ""
            else:
                value = _content_value(cell.value)
            values.append(_md_cell(value))
        lines.append("| " + " | ".join(values) + " |")
        if row == min_row:
            lines.append("| " + " | ".join("---" for _ in range(min_col, max_col + 1)) + " |")
    return "\n".join(lines)


def _merged_non_owner(sheet: Any, coordinate: str) -> bool:
    for merged in sheet.merged_cells.ranges:
        if coordinate not in merged:
            continue
        return coordinate != merged.start_cell.coordinate
    return False


def _render_format_table(
    sheet: Any,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
    warnings: list[str],
    image_map: dict[str, str],
    assets_dir: Path,
) -> str:
    lines = [f"# {sheet.title}", ""]
    merge_names = {str(item): f"merge{index}" for index, item in enumerate(sorted(sheet.merged_cells.ranges, key=str), 1)}
    for row in range(min_row, max_row + 1):
        cells = []
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row, col)
            directives: list[str] = []
            for merged, name in merge_names.items():
                if _coordinate_in_range(cell.row, cell.column, merged):
                    directives.append(f"merge:{name}")
                    break
            directives.extend(_style_directives(cell, sheet, row, col, min_row, min_col, warnings, image_map, assets_dir))
            cells.append(_md_cell("; ".join(directives)))
        lines.append("| " + " | ".join(cells) + " |")
        if row == min_row:
            lines.append("| " + " | ".join("---" for _ in range(min_col, max_col + 1)) + " |")
    return "\n".join(lines)


def _style_directives(
    cell: Any,
    sheet: Any,
    row: int,
    col: int,
    min_row: int,
    min_col: int,
    warnings: list[str],
    image_map: dict[str, str],
    assets_dir: Path,
) -> list[str]:
    directives: list[str] = []
    font = cell.font
    if font.bold:
        directives.append("font-weight:bold")
    if font.italic:
        directives.append("font-style:italic")
    if font.underline:
        directives.append("text-decoration:underline")
    if font.sz:
        directives.append(f"font-size:{_number(font.sz)}")
    if font.name:
        directives.append(f"font-family:{_safe_value(font.name)}")
    font_color = _color(font.color)
    if font_color:
        directives.append(f"color:{font_color}")
    fill_color = _color(cell.fill.fgColor) if cell.fill.fill_type else None
    if fill_color:
        directives.append(f"background:{fill_color}")
    if cell.alignment.horizontal:
        directives.append(f"text-align:{cell.alignment.horizontal}")
    if cell.alignment.vertical:
        directives.append(f"vertical-align:{cell.alignment.vertical}")
    if cell.alignment.wrap_text:
        directives.append("wrap")
    if cell.number_format and cell.number_format != "General":
        directives.append(f"number-format:{_safe_value(cell.number_format)}")
    _append_border_directives(cell, directives)
    if cell.hyperlink and cell.hyperlink.target:
        directives.append(f"hyperlink:{_safe_value(cell.hyperlink.target)}")
    if cell.comment and cell.comment.text:
        directives.append(f"comment:{_safe_value(cell.comment.text)}")
    image_path = image_map.get(cell.coordinate)
    if image_path:
        directives.append(f"image:{_safe_value(image_path)}")
    if row == min_row and col == min_col:
        if sheet.freeze_panes:
            directives.append(f"freeze:{sheet.freeze_panes}")
        if sheet.auto_filter.ref:
            directives.append(f"filter:{sheet.auto_filter.ref}")
        if sheet.sheet_properties.tabColor and sheet.sheet_properties.tabColor.rgb:
            directives.append(f"tab-color:{_color(sheet.sheet_properties.tabColor)}")
        if sheet.page_setup.orientation:
            directives.append(f"page-orientation:{sheet.page_setup.orientation}")
    width = sheet.column_dimensions[get_column_letter(col)].width
    if width is not None and row == min_row:
        directives.append(f"width:{_number(width)}")
    height = sheet.row_dimensions[row].height
    if height is not None:
        directives.append(f"height:{_number(height)}")
    _append_table_directives(sheet, directives, row, col, min_row, min_col)
    _append_validation_directives(sheet, cell.coordinate, directives, warnings)
    if row == min_row and col == min_col:
        _append_conditional_directives(sheet, directives, warnings)
    return directives


def _append_border_directives(cell: Any, directives: list[str]) -> None:
    sides = [cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]
    styles = [side.style for side in sides]
    if styles and styles.count(styles[0]) == len(styles) and styles[0]:
        directives.append(f"border:{styles[0]}")
        colour = _color(sides[0].color)
        if colour:
            directives.append(f"border-color:{colour}")
        return
    for name, side in (("left", cell.border.left), ("right", cell.border.right), ("top", cell.border.top), ("bottom", cell.border.bottom)):
        if side.style:
            directives.append(f"border-{name}:{side.style}")


def _append_table_directives(sheet: Any, directives: list[str], row: int, col: int, min_row: int, min_col: int) -> None:
    if row != min_row or col != min_col:
        return
    for table in sheet.tables.values():
        if table.ref:
            directives.append("table:on")
            if table.tableStyleInfo and table.tableStyleInfo.name:
                directives.append(f"table-style:{table.tableStyleInfo.name}")


def _append_validation_directives(sheet: Any, coordinate: str, directives: list[str], warnings: list[str]) -> None:
    for validation in sheet.data_validations.dataValidation:
        if not any(coordinate in str(item) for item in validation.sqref.ranges):
            continue
        formula = validation.formula1
        if validation.type == "list" and isinstance(formula, str) and formula.startswith('"') and formula.endswith('"'):
            choices = formula[1:-1].replace('""', '"')
            directives.append(f"validation:{_safe_value(choices)}@{validation.sqref}")
        else:
            warnings.append(f"{sheet.title}!{coordinate}: 数据验证 {validation.type} 无法完整映射为当前 Markdown 规则。")


def _append_conditional_directives(sheet: Any, directives: list[str], warnings: list[str]) -> None:
    for conditional in sheet.conditional_formatting:
        target = str(conditional.sqref)
        for rule in sheet.conditional_formatting[conditional]:
            if rule.type == "dataBar" and rule.dataBar is not None:
                colour = _color(rule.dataBar.color)
                if colour:
                    directives.append(f"data-bar:{colour}@{target}")
                else:
                    warnings.append(f"{sheet.title}: data bar 缺少可转换颜色。")
            elif rule.type == "colorScale" and rule.colorScale is not None:
                colours = [_color(item) for item in getattr(rule.colorScale, "color", [])]
                if len(colours) == 3 and all(colours):
                    directives.append(f"color-scale:{','.join(colours)}@{target}")
                else:
                    warnings.append(f"{sheet.title}: color scale 不是当前规则支持的三色形式。")
            elif rule.type == "cellIs":
                fill = _rule_fill_color(rule)
                formula = (rule.formula or [""])[0]
                if fill and rule.operator and formula:
                    directives.append(f"cell-is:{rule.operator}{formula}\\|{fill}@{target}")
                else:
                    warnings.append(f"{sheet.title}: cell-is 条件格式无法完整映射。")
            elif rule.type == "expression":
                fill = _rule_fill_color(rule)
                formula = (rule.formula or [""])[0]
                if fill and formula:
                    directives.append(f"formula-rule:{formula}\\|{fill}@{target}")
                else:
                    warnings.append(f"{sheet.title}: formula 条件格式无法完整映射。")
            else:
                warnings.append(f"{sheet.title}: 条件格式类型 {rule.type} 无法映射为当前 Markdown 规则。")


def _rule_fill_color(rule: Any) -> str | None:
    dxf = getattr(rule, "dxf", None)
    fill = getattr(dxf, "fill", None) if dxf else None
    return _color(fill.fgColor) if fill and fill.fill_type else None


def _coordinate_in_range(row: int, column: int, reference: str) -> bool:
    min_col, min_row, max_col, max_row = range_boundaries(reference)
    return min_row <= row <= max_row and min_col <= column <= max_col


def _export_images(sheet: Any, assets_dir: Path, warnings: list[str]) -> dict[str, str]:
    exported: dict[str, str] = {}
    for index, image in enumerate(getattr(sheet, "_images", []) or [], 1):
        try:
            anchor = image.anchor._from
            target_dir = assets_dir / _safe_filename(sheet.title)
            target_dir.mkdir(parents=True, exist_ok=True)
            suffix = "." + (getattr(image, "format", None) or "png").lower()
            target = target_dir / f"{get_column_letter(anchor.col + 1)}{anchor.row + 1}_{index}{suffix}"
            target.write_bytes(image._data())
            relative = str(target.relative_to(assets_dir)).replace("\\", "/")
            exported[f"{get_column_letter(anchor.col + 1)}{anchor.row + 1}"] = f"{assets_dir.name}/{relative}"
        except Exception as exc:
            warnings.append(f"{sheet.title}: 图片提取失败：{exc}")
    return exported


def _collect_unsupported_objects(sheet: Any, warnings: list[str]) -> None:
    if getattr(sheet, "_charts", None):
        warnings.append(f"{sheet.title}: 存在图表，当前反向规则暂未生成 chart 指令；图表未写入 format.md。")
    if getattr(sheet, "_images", None):
        return


def _render_report(input_path: Path, content_path: Path, format_path: Path, sheets: list[Any], warnings: list[str], assets: list[str]) -> str:
    lines = [
        "# Excel 反向提取报告",
        "",
        f"- 输入文件：`{input_path}`",
        f"- 内容文件：`{content_path}`",
        f"- 格式文件：`{format_path}`",
        f"- Sheet 数量：{len(sheets)}",
        f"- 图片数量：{len(assets)}",
        "",
        "## 给 AI 的读取规则",
        "",
        "1. 先读取 content.md，理解表格内容和公式。",
        "2. 再读取 format.md；它与 content.md 按标题和单元格位置一一对应。",
        "3. 内容表只修改值，格式表只修改格式和结构指令。",
        "4. 仅在需要重新生成 Excel 时使用 markdown_to_excel；若当前没有该工具，先从在线市场安装，不能自行猜测单元格偏移。",
        "5. 本次提取是规范化结果，不保证还原原始 Excel 编辑过程。",
    ]
    if warnings:
        lines.extend(["", "## 未完整映射项目", "", *[f"- {item}" for item in warnings]])
    return "\n".join(lines) + "\n"


def _content_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if hasattr(value, "date") and value.__class__.__name__ == "datetime":
        return f"date:{value.date().isoformat()}"
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return f"date:{value.isoformat()}"
    return str(value)


def _md_cell(value: str) -> str:
    return value.replace("\\", "\\\\").replace("|", "\\|").replace("\n", "<br>")


def _safe_value(value: Any) -> str:
    return str(value).replace(";", ",").replace("\n", " ").replace("\r", " ")


def _safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("._") or "sheet"


def _number(value: Any) -> str:
    number = float(value)
    return str(int(number)) if number.is_integer() else f"{number:.4f}".rstrip("0").rstrip(".")


def _color(value: Any) -> str | None:
    if value is None:
        return None
    rgb = getattr(value, "rgb", None)
    if isinstance(rgb, str):
        raw = rgb[-6:]
        if len(raw) == 6 and raw.upper() not in {"000000", "FFFFFF"}:
            return f"#{raw.upper()}"
        if raw.upper() == "FFFFFF":
            return "#FFFFFF"
        if raw.upper() == "000000":
            return "#000000"
    return None
