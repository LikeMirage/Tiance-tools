from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_charts import add_chart
from excel_errors import ToolError
from excel_formulas import formula_value
from excel_rules import add_conditional_format
from excel_styles import apply_style, apply_style_to_range


def edit_workbook(
    payload: dict[str, Any],
    *,
    root: Path,
    validate_only: bool = False,
) -> dict[str, Any]:
    input_path = _resolve_input_path(payload, root)
    output_path = _resolve_output_path(payload, root, input_path)
    overwrite = _read_bool(payload.get("overwrite"), False)
    if output_path == input_path and not overwrite:
        raise ToolError(
            "UNSAFE_OVERWRITE",
            "编辑已有 Excel 默认不能覆盖原文件，请设置 output_path 另存为新文件。",
            {"input_path": str(input_path)},
        )
    if output_path.exists() and not overwrite and not _is_implicit_output(payload):
        raise ToolError("OUTPUT_EXISTS", "输出文件已存在。", {"output_path": str(output_path)})

    operations = payload.get("operations")
    if not isinstance(operations, list) or not operations:
        raise ToolError("INVALID_ARGUMENT", "edit 操作必须提供非空 operations 数组。")

    workbook = load_workbook(input_path)
    try:
        summaries: list[dict[str, Any]] = []
        warnings: list[str] = []
        for index, operation in enumerate(operations):
            if not isinstance(operation, dict):
                raise ToolError(
                    "INVALID_OPERATION",
                    "operations 的每一项都必须是对象。",
                    {"operation_index": index},
                )
            summaries.append(_apply_operation(workbook, operation, index, warnings))

        if validate_only:
            return {
                "input_path": str(input_path),
                "output_path": str(output_path),
                "operation_count": len(summaries),
                "operations": summaries,
                "warnings": warnings,
            }

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        _validate_saved_workbook(output_path)
        return {
            "input_path": str(input_path),
            "output_path": str(output_path),
            "operation_count": len(summaries),
            "operations": summaries,
            "overwrite": overwrite,
            "warnings": warnings,
        }
    finally:
        workbook.close()


def _apply_operation(workbook: Any, operation: dict[str, Any], index: int, warnings: list[str]) -> dict[str, Any]:
    operation_type = str(operation.get("type") or "").strip()
    if not operation_type:
        raise _operation_error(index, operation_type, "INVALID_OPERATION", "operation.type 不能为空。")

    try:
        if operation_type == "set_cells":
            return _set_cells(workbook, operation)
        if operation_type == "append_rows":
            return _append_rows(workbook, operation)
        if operation_type == "write_block":
            return _write_block(workbook, operation)
        if operation_type == "add_sheet":
            return _add_sheet(workbook, operation)
        if operation_type == "delete_sheet":
            return _delete_sheet(workbook, operation)
        if operation_type == "rename_sheet":
            return _rename_sheet(workbook, operation)
        if operation_type == "copy_sheet":
            return _copy_sheet(workbook, operation)
        if operation_type == "merge_cells":
            return _merge_cells(workbook, operation)
        if operation_type == "unmerge_cells":
            return _unmerge_cells(workbook, operation, warnings)
        if operation_type == "set_style":
            return _set_style(workbook, operation)
        if operation_type == "set_dimensions":
            return _set_dimensions(workbook, operation)
        if operation_type == "insert_rows":
            return _insert_rows(workbook, operation)
        if operation_type == "delete_rows":
            return _delete_rows(workbook, operation)
        if operation_type == "insert_columns":
            return _insert_columns(workbook, operation)
        if operation_type == "delete_columns":
            return _delete_columns(workbook, operation)
        if operation_type == "clear_range":
            return _clear_range(workbook, operation)
        if operation_type == "set_freeze_panes":
            return _set_freeze_panes(workbook, operation)
        if operation_type == "set_auto_filter":
            return _set_auto_filter(workbook, operation)
        if operation_type == "add_table":
            return _add_table(workbook, operation)
        if operation_type == "remove_table":
            return _remove_table(workbook, operation)
        if operation_type == "add_data_validation":
            return _add_data_validation(workbook, operation)
        if operation_type == "add_conditional_format":
            return _add_conditional_format(workbook, operation)
        if operation_type == "add_chart":
            return _add_chart(workbook, operation)
        if operation_type == "set_page":
            return _set_page(workbook, operation)
    except ToolError as exc:
        details = {"operation_index": index, "operation_type": operation_type, **exc.details}
        raise ToolError(exc.code, exc.message, details) from exc

    raise _operation_error(
        index,
        operation_type,
        "UNSUPPORTED_OPERATION",
        "不支持的 Excel 编辑操作。",
    )


def _set_cells(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    cells = operation.get("cells")
    if not isinstance(cells, list) or not cells:
        raise ToolError("INVALID_ARGUMENT", "set_cells.cells 必须是非空数组。")
    count = 0
    for spec in cells:
        if not isinstance(spec, dict):
            raise ToolError("INVALID_ARGUMENT", "set_cells.cells 每项都必须是对象。")
        coordinate = _required_text(spec, "cell")
        cell = sheet[coordinate]
        if _read_bool(spec.get("clear"), False):
            cell.value = None
        elif isinstance(spec.get("formula"), str):
            cell.value = formula_value(spec["formula"], f"{sheet.title}!{coordinate}")
        elif "value" in spec:
            cell.value = spec["value"]
        apply_style(cell, spec.get("style"))
        if isinstance(spec.get("number_format"), str):
            cell.number_format = spec["number_format"]
        if isinstance(spec.get("hyperlink"), str):
            cell.hyperlink = spec["hyperlink"]
            if cell.value is None:
                cell.value = spec["hyperlink"]
        if isinstance(spec.get("comment"), str):
            cell.comment = Comment(spec["comment"], "天策")
        count += 1
    return {"type": "set_cells", "sheet": sheet.title, "cells": count}


def _append_rows(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    rows = operation.get("rows")
    if not isinstance(rows, list) or not rows:
        raise ToolError("INVALID_ARGUMENT", "append_rows.rows 必须是非空二维数组。")
    for row in rows:
        if not isinstance(row, list):
            raise ToolError("INVALID_ARGUMENT", "append_rows.rows 的每一行都必须是数组。")
        sheet.append([_cell_item_value(item, sheet.title) for item in row])
    return {"type": "append_rows", "sheet": sheet.title, "rows": len(rows)}


def _write_block(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    start_cell = str(operation.get("start_cell") or "A1")
    start_row, start_col = coordinate_to_tuple(start_cell)
    rows = operation.get("rows")
    if not isinstance(rows, list) or not rows:
        raise ToolError("INVALID_ARGUMENT", "write_block.rows 必须是非空二维数组。")
    count = 0
    for row_offset, row in enumerate(rows):
        if not isinstance(row, list):
            raise ToolError("INVALID_ARGUMENT", "write_block.rows 的每一行都必须是数组。")
        for col_offset, value in enumerate(row):
            cell = sheet.cell(row=start_row + row_offset, column=start_col + col_offset)
            cell.value = _cell_item_value(value, sheet.title, cell.coordinate)
            apply_style(cell, operation.get("style"))
            count += 1
    return {"type": "write_block", "sheet": sheet.title, "cells": count}


def _add_sheet(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    name = _required_text(operation, "name")
    if name in workbook.sheetnames:
        raise ToolError("DUPLICATE_SHEET_NAME", "Sheet 名称不能重复。", {"name": name})
    index = operation.get("index") if isinstance(operation.get("index"), int) else None
    sheet = workbook.create_sheet(name, index=index)
    return {"type": "add_sheet", "sheet": sheet.title}


def _delete_sheet(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    if len(workbook.worksheets) <= 1:
        raise ToolError("INVALID_OPERATION", "不能删除工作簿里的最后一个 Sheet。")
    workbook.remove(sheet)
    return {"type": "delete_sheet", "sheet": sheet.title}


def _rename_sheet(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    new_name = _required_text(operation, "name")
    old_name = sheet.title
    sheet.title = new_name
    return {"type": "rename_sheet", "from": old_name, "to": sheet.title}


def _copy_sheet(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    source = _sheet(workbook, operation)
    target = workbook.copy_worksheet(source)
    if isinstance(operation.get("name"), str) and operation["name"].strip():
        target.title = operation["name"].strip()
    return {"type": "copy_sheet", "from": source.title, "to": target.title}


def _merge_cells(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    ranges = _ranges(operation)
    for target_range in ranges:
        sheet.merge_cells(target_range)
    return {"type": "merge_cells", "sheet": sheet.title, "ranges": len(ranges)}


def _unmerge_cells(workbook: Any, operation: dict[str, Any], warnings: list[str]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    ranges = _ranges(operation)
    count = 0
    for target_range in ranges:
        try:
            sheet.unmerge_cells(target_range)
            count += 1
        except ValueError:
            warnings.append(f"{sheet.title}: {target_range} 不是已合并区域，已跳过。")
    return {"type": "unmerge_cells", "sheet": sheet.title, "ranges": count}


def _set_style(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    target_range = _target_range(operation)
    count = apply_style_to_range(
        sheet,
        target_range,
        operation.get("style"),
        operation.get("number_format") if isinstance(operation.get("number_format"), str) else None,
    )
    return {"type": "set_style", "sheet": sheet.title, "range": target_range, "cells": count}


def _set_dimensions(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    columns = 0
    rows = 0
    column_widths = operation.get("column_widths")
    if isinstance(column_widths, dict):
        for key, value in column_widths.items():
            if not isinstance(value, (int, float)):
                continue
            for column in _expand_column_key(str(key)):
                sheet.column_dimensions[column].width = value
                columns += 1
    row_heights = operation.get("row_heights")
    if isinstance(row_heights, dict):
        for key, value in row_heights.items():
            if isinstance(value, (int, float)) and str(key).isdigit():
                sheet.row_dimensions[int(key)].height = value
                rows += 1
    return {"type": "set_dimensions", "sheet": sheet.title, "columns": columns, "rows": rows}


def _insert_rows(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    index = _positive_int(operation, "index")
    count = _positive_int(operation, "count", default=1)
    sheet.insert_rows(index, count)
    return {"type": "insert_rows", "sheet": sheet.title, "index": index, "count": count}


def _delete_rows(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    index = _positive_int(operation, "index")
    count = _positive_int(operation, "count", default=1)
    sheet.delete_rows(index, count)
    return {"type": "delete_rows", "sheet": sheet.title, "index": index, "count": count}


def _insert_columns(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    index = _positive_int(operation, "index")
    count = _positive_int(operation, "count", default=1)
    sheet.insert_cols(index, count)
    return {"type": "insert_columns", "sheet": sheet.title, "index": index, "count": count}


def _delete_columns(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    index = _positive_int(operation, "index")
    count = _positive_int(operation, "count", default=1)
    sheet.delete_cols(index, count)
    return {"type": "delete_columns", "sheet": sheet.title, "index": index, "count": count}


def _clear_range(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    target_range = _target_range(operation)
    count = 0
    for row in sheet[target_range]:
        for cell in row:
            cell.value = None
            count += 1
    return {"type": "clear_range", "sheet": sheet.title, "range": target_range, "cells": count}


def _set_freeze_panes(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    sheet.freeze_panes = operation.get("cell") if isinstance(operation.get("cell"), str) else None
    return {"type": "set_freeze_panes", "sheet": sheet.title, "cell": sheet.freeze_panes}


def _set_auto_filter(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    sheet.auto_filter.ref = operation.get("range") if isinstance(operation.get("range"), str) else None
    return {"type": "set_auto_filter", "sheet": sheet.title, "range": sheet.auto_filter.ref}


def _add_table(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    name = _required_text(operation, "name")
    target_range = _required_text(operation, "range")
    table = Table(displayName=name, ref=target_range)
    table.tableStyleInfo = TableStyleInfo(
        name=str(operation.get("style") or "TableStyleMedium9"),
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=_read_bool(operation.get("show_row_stripes"), True),
        showColumnStripes=_read_bool(operation.get("show_column_stripes"), False),
    )
    sheet.add_table(table)
    return {"type": "add_table", "sheet": sheet.title, "name": name, "range": target_range}


def _remove_table(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    name = _required_text(operation, "name")
    if name not in sheet.tables:
        raise ToolError("TABLE_NOT_FOUND", "指定表格不存在。", {"sheet": sheet.title, "name": name})
    del sheet.tables[name]
    return {"type": "remove_table", "sheet": sheet.title, "name": name}


def _add_data_validation(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    target_range = _required_text(operation, "range")
    validation_type = _required_text(operation, "validation_type")
    validation = DataValidation(
        type=validation_type,
        operator=operation.get("operator"),
        formula1=operation.get("formula1"),
        formula2=operation.get("formula2"),
        allow_blank=_read_bool(operation.get("allow_blank"), True),
    )
    if isinstance(operation.get("prompt"), str):
        validation.prompt = operation["prompt"]
        validation.showInputMessage = True
    if isinstance(operation.get("error"), str):
        validation.error = operation["error"]
        validation.showErrorMessage = True
    validation.add(target_range)
    sheet.add_data_validation(validation)
    return {"type": "add_data_validation", "sheet": sheet.title, "range": target_range}


def _add_conditional_format(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    format_type = _required_text(operation, "format_type")
    spec = {**operation, "type": format_type}
    add_conditional_format(sheet, spec)
    return {"type": "add_conditional_format", "sheet": sheet.title, "range": operation.get("range")}


def _add_chart(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    chart_type = _required_text(operation, "chart_type")
    spec = {**operation, "type": chart_type}
    add_chart(sheet, spec)
    return {"type": "add_chart", "sheet": sheet.title, "chart_type": chart_type}


def _set_page(workbook: Any, operation: dict[str, Any]) -> dict[str, Any]:
    sheet = _sheet(workbook, operation)
    if isinstance(operation.get("orientation"), str):
        sheet.page_setup.orientation = operation["orientation"]
    if isinstance(operation.get("paper_size"), int):
        sheet.page_setup.paperSize = operation["paper_size"]
    if isinstance(operation.get("fit_to_width"), int):
        sheet.page_setup.fitToWidth = operation["fit_to_width"]
    if isinstance(operation.get("fit_to_height"), int):
        sheet.page_setup.fitToHeight = operation["fit_to_height"]
    if operation.get("fit_to_width") is not None or operation.get("fit_to_height") is not None:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
    return {"type": "set_page", "sheet": sheet.title}


def _cell_item_value(item: Any, sheet_name: str, coordinate: str | None = None) -> Any:
    location = f"{sheet_name}!{coordinate}" if coordinate else sheet_name
    if isinstance(item, dict):
        if isinstance(item.get("formula"), str):
            return formula_value(item["formula"], location)
        if "value" in item:
            return item["value"]
    if isinstance(item, str) and item.strip().startswith("="):
        return formula_value(item, location)
    return item


def _sheet(workbook: Any, operation: dict[str, Any]) -> Any:
    name = _required_text(operation, "sheet")
    if name not in workbook.sheetnames:
        raise ToolError("SHEET_NOT_FOUND", "指定 Sheet 不存在。", {"sheet": name})
    return workbook[name]


def _ranges(operation: dict[str, Any]) -> list[str]:
    ranges = operation.get("ranges")
    if isinstance(ranges, list):
        cleaned = [str(item).strip() for item in ranges if str(item).strip()]
        if cleaned:
            return cleaned
    target_range = operation.get("range")
    if isinstance(target_range, str) and target_range.strip():
        return [target_range.strip()]
    raise ToolError("INVALID_ARGUMENT", "操作必须提供 range 或 ranges。")


def _target_range(operation: dict[str, Any]) -> str:
    if isinstance(operation.get("range"), str) and operation["range"].strip():
        return operation["range"].strip()
    if isinstance(operation.get("cell"), str) and operation["cell"].strip():
        return operation["cell"].strip()
    raise ToolError("INVALID_ARGUMENT", "操作必须提供 range 或 cell。")


def _required_text(payload: dict[str, Any], key: str) -> str:
    value = payload.get(key)
    if not isinstance(value, str) or not value.strip():
        raise ToolError("INVALID_ARGUMENT", f"{key} 必须是非空字符串。")
    return value.strip()


def _positive_int(payload: dict[str, Any], key: str, default: int | None = None) -> int:
    value = payload.get(key, default)
    if not isinstance(value, int) or value < 1:
        raise ToolError("INVALID_ARGUMENT", f"{key} 必须是正整数。")
    return value


def _read_bool(value: Any, default: bool = False) -> bool:
    return value if isinstance(value, bool) else default


def _expand_column_key(key: str) -> list[str]:
    cleaned = key.strip().upper()
    if ":" not in cleaned:
        return [cleaned]
    start, end = cleaned.split(":", 1)
    start_index = _column_index(start)
    end_index = _column_index(end)
    if end_index < start_index:
        start_index, end_index = end_index, start_index
    return [get_column_letter(index) for index in range(start_index, end_index + 1)]


def _column_index(column: str) -> int:
    total = 0
    for char in column:
        if not ("A" <= char <= "Z"):
            raise ToolError("INVALID_ARGUMENT", "列宽键必须是列字母或列范围。", {"column": column})
        total = total * 26 + (ord(char) - ord("A") + 1)
    return total


def _resolve_input_path(payload: dict[str, Any], root: Path) -> Path:
    raw = payload.get("input_path")
    if not isinstance(raw, str) or not raw.strip():
        raise ToolError("INVALID_ARGUMENT", "input_path 必须是非空字符串。")
    path = Path(raw.strip()).expanduser()
    if not path.is_absolute():
        path = root / path
    resolved = path.resolve(strict=False)
    _ensure_inside_workspace(resolved, root, "input_path")
    if not resolved.is_file():
        raise ToolError("INPUT_NOT_FOUND", "input_path 指向的 Excel 文件不存在。", {"input_path": str(resolved)})
    if resolved.suffix.lower() != ".xlsx":
        raise ToolError("INVALID_ARGUMENT", "目前只支持编辑 .xlsx 文件。", {"input_path": str(resolved)})
    return resolved


def _resolve_output_path(payload: dict[str, Any], root: Path, input_path: Path) -> Path:
    raw = payload.get("output_path")
    if isinstance(raw, str) and raw.strip():
        path = Path(raw.strip()).expanduser()
        if not path.is_absolute():
            path = root / path
        resolved = path.resolve(strict=False)
        _ensure_inside_workspace(resolved, root, "output_path")
        return resolved if resolved.suffix.lower() == ".xlsx" else resolved.with_suffix(".xlsx")
    return _unique_edited_path(input_path)


def _unique_edited_path(input_path: Path) -> Path:
    base = input_path.with_name(f"{input_path.stem}_edited.xlsx")
    if not base.exists():
        return base
    index = 2
    while True:
        candidate = input_path.with_name(f"{input_path.stem}_edited_{index}.xlsx")
        if not candidate.exists():
            return candidate
        index += 1


def _is_implicit_output(payload: dict[str, Any]) -> bool:
    return not (isinstance(payload.get("output_path"), str) and payload["output_path"].strip())


def _ensure_inside_workspace(path: Path, root: Path, field: str) -> None:
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise ToolError(
            "PATH_OUTSIDE_WORKSPACE",
            f"{field} 必须位于工作区内。",
            {field: str(path), "workspace_root": str(root)},
        ) from exc


def _operation_error(index: int, operation_type: str, code: str, message: str) -> ToolError:
    return ToolError(code, message, {"operation_index": index, "operation_type": operation_type})


def _validate_saved_workbook(path: Path) -> None:
    try:
        workbook = load_workbook(path, data_only=False, read_only=True)
        workbook.close()
    except Exception as exc:
        raise ToolError(
            "OUTPUT_INVALID",
            "保存后的 Excel 文件无法重新读取，已阻止输出坏文件。",
            {"output_path": str(path), "reason": str(exc) or type(exc).__name__},
        ) from exc
