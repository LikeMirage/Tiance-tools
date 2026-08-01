from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def normalize_color(value: Any) -> str | None:
    if value is None:
        return None
    color = str(value).strip().lstrip("#")
    if len(color) in {6, 8} and all(char in "0123456789abcdefABCDEF" for char in color):
        return color.upper()
    return None


def build_font(spec: Any) -> Font | None:
    if not isinstance(spec, dict):
        return None
    kwargs: dict[str, Any] = {}
    for key in ("name", "size", "bold", "italic", "underline", "strike"):
        if key in spec:
            kwargs[key] = spec[key]
    color = normalize_color(spec.get("color"))
    if color:
        kwargs["color"] = color
    return Font(**kwargs)


def build_fill(spec: Any) -> PatternFill | None:
    if not isinstance(spec, dict):
        return None
    color = normalize_color(spec.get("color") or spec.get("fgColor") or spec.get("fg_color"))
    if not color:
        return None
    fill_type = str(spec.get("fill_type") or spec.get("pattern_type") or "solid")
    return PatternFill(fill_type=fill_type, fgColor=color)


def build_alignment(spec: Any) -> Alignment | None:
    if not isinstance(spec, dict):
        return None
    allowed = {
        "horizontal",
        "vertical",
        "text_rotation",
        "wrap_text",
        "shrink_to_fit",
        "indent",
    }
    kwargs = {key: spec[key] for key in allowed if key in spec}
    return Alignment(**kwargs) if kwargs else None


def build_border(spec: Any) -> Border | None:
    if not isinstance(spec, dict):
        return None
    side_names = ("left", "right", "top", "bottom")
    if any(name in spec and isinstance(spec[name], dict) for name in side_names):
        sides = {name: _build_side(spec.get(name)) for name in side_names}
        return Border(**sides)
    side = _build_side(spec)
    return Border(left=side, right=side, top=side, bottom=side)


def apply_style(cell: Any, style: Any) -> None:
    if not isinstance(style, dict):
        return
    font = build_font(style.get("font"))
    fill = build_fill(style.get("fill"))
    alignment = build_alignment(style.get("alignment"))
    border = build_border(style.get("border"))
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if isinstance(style.get("number_format"), str):
        cell.number_format = style["number_format"]


def apply_style_to_range(ws: Any, range_text: str, style: Any, number_format: str | None = None) -> int:
    count = 0
    for row in ws[range_text]:
        for cell in row:
            apply_style(cell, style)
            if number_format:
                cell.number_format = number_format
            count += 1
    return count


def _build_side(spec: Any) -> Side:
    if not isinstance(spec, dict):
        return Side(style="thin", color="D9D9D9")
    style = str(spec.get("style") or "thin")
    color = normalize_color(spec.get("color")) or "D9D9D9"
    return Side(style=style, color=color)
