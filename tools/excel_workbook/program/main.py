from __future__ import annotations

import json
import os
from pathlib import Path
import re
import sys
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from tiance_runtime import run_tool
from excel_charts import add_chart
from excel_edit import edit_workbook
from excel_errors import ToolError
from excel_formulas import formula_value
from excel_inspect import inspect_workbook
from excel_markdown_export import export_workbook
from excel_rules import add_conditional_format
from excel_styles import apply_style, apply_style_to_range, normalize_color


INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def ok(summary: str, data: dict[str, Any], warnings: list[str] | None = None) -> dict[str, Any]:
    return {"ok": True, "summary": summary, "data": data, "warnings": warnings or []}


def fail(
    code: str,
    message: str,
    details: dict[str, Any] | None = None,
    warnings: list[str] | None = None,
) -> dict[str, Any]:
    return {
        "ok": False,
        "error": f"{code}: {message}",
        "error_info": {"code": code, "message": message, "details": details or {}},
        "warnings": warnings or [],
    }


def workspace_root() -> Path:
    raw = os.environ.get("TIANCE_WORKSPACE_ROOT") or os.environ.get("WORKSPACE_ROOT") or os.getcwd()
    return Path(raw).expanduser().resolve(strict=False)


def read_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "1", "yes", "y", "on"}:
            return True
        if lowered in {"false", "0", "no", "n", "off"}:
            return False
    return default


def resolve_output_path(payload: dict[str, Any], root: Path) -> Path:
    raw = payload.get("output_path")
    if not isinstance(raw, str) or not raw.strip():
        raise ToolError("INVALID_ARGUMENT", "output_path 必须是非空字符串。")
    path = Path(raw.strip()).expanduser()
    if not path.is_absolute():
        path = root / path
    resolved = path.resolve(strict=False)
    try:
        resolved.relative_to(root)
    except ValueError as exc:
        raise ToolError(
            "PATH_OUTSIDE_WORKSPACE",
            "output_path 必须位于工作区内。",
            {"output_path": str(resolved), "workspace_root": str(root)},
        ) from exc
    if resolved.suffix.lower() != ".xlsx":
        resolved = resolved.with_suffix(".xlsx")
    return resolved


def resolve_workspace_path(raw: Any, root: Path, *, label: str, suffix: str | None = None) -> Path:
    if not isinstance(raw, str) or not raw.strip():
        raise ToolError("INVALID_ARGUMENT", f"{label} 必须是非空字符串。")
    path = Path(raw.strip()).expanduser()
    if not path.is_absolute():
        path = root / path
    resolved = path.resolve(strict=False)
    try:
        resolved.relative_to(root)
    except ValueError as exc:
        raise ToolError("PATH_OUTSIDE_WORKSPACE", f"{label} 必须位于工作区内。") from exc
    if suffix and resolved.suffix.lower() != suffix:
        resolved = resolved.with_suffix(suffix)
    return resolved


def apply_workbook_properties(wb: Workbook, value: Any) -> None:
    if not isinstance(value, dict):
        return
    allowed = {
        "title",
        "subject",
        "creator",
        "keywords",
        "description",
        "category",
        "company",
        "manager",
        "lastModifiedBy",
    }
    for key in allowed:
        item = value.get(key)
        if isinstance(item, str):
            setattr(wb.properties, key, item)


def validate_sheet_name(name: Any, used: set[str]) -> str:
    if not isinstance(name, str) or not name.strip():
        raise ToolError("INVALID_ARGUMENT", "每个 sheet.name 都必须是非空字符串。")
    normalized = name.strip()
    if len(normalized) > 31:
        raise ToolError("INVALID_SHEET_NAME", "Sheet 名称不能超过 31 个字符。", {"name": normalized})
    if INVALID_SHEET_CHARS.search(normalized):
        raise ToolError("INVALID_SHEET_NAME", "Sheet 名称包含 Excel 禁止字符。", {"name": normalized})
    if normalized in used:
        raise ToolError("DUPLICATE_SHEET_NAME", "Sheet 名称不能重复。", {"name": normalized})
    used.add(normalized)
    return normalized


def write_data_block(ws: Any, block: dict[str, Any]) -> int:
    start_cell = str(block.get("start_cell") or "A1")
    start_row, start_col = coordinate_to_tuple(start_cell)
    rows = block.get("rows")
    if not isinstance(rows, list):
        raise ToolError("INVALID_ARGUMENT", "data_blocks.rows 必须是二维数组。")
    count = 0
    for row_offset, row_values in enumerate(rows):
        if not isinstance(row_values, list):
            raise ToolError("INVALID_ARGUMENT", "data_blocks.rows 的每一行都必须是数组。")
        for col_offset, item in enumerate(row_values):
            cell = ws.cell(row=start_row + row_offset, column=start_col + col_offset)
            set_cell_value(cell, item)
            apply_style(cell, block.get("style"))
            if row_offset == 0 and read_bool(block.get("header"), False):
                apply_style(cell, block.get("header_style"))
            count += 1
    return count


def set_cell_value(cell: Any, item: Any) -> None:
    if isinstance(item, dict):
        if isinstance(item.get("formula"), str):
            cell.value = _formula(item["formula"], cell.coordinate)
        elif "value" in item:
            cell.value = item["value"]
        else:
            cell.value = json.dumps(item, ensure_ascii=False)
        apply_style(cell, item.get("style"))
        if isinstance(item.get("number_format"), str):
            cell.number_format = item["number_format"]
        return
    if isinstance(item, str) and item.strip().startswith("="):
        cell.value = _formula(item, cell.coordinate)
        return
    cell.value = item


def write_cell_spec(ws: Any, spec: dict[str, Any]) -> int:
    coordinate = str(spec.get("cell") or "").strip()
    if not coordinate:
        raise ToolError("INVALID_ARGUMENT", "cells 每项都必须提供 cell。")
    cell = ws[coordinate]
    if isinstance(spec.get("formula"), str):
        cell.value = _formula(spec["formula"], coordinate)
    elif "value" in spec:
        cell.value = spec["value"]
    apply_style(cell, spec.get("style"))
    if isinstance(spec.get("number_format"), str):
        cell.number_format = spec["number_format"]
    if isinstance(spec.get("hyperlink"), str):
        cell.hyperlink = spec["hyperlink"]
        if cell.value is None:
            cell.value = spec["hyperlink"]
    if isinstance(spec.get("comment"), str):
        cell.comment = Comment(spec["comment"], "天策")
    return 1


def apply_dimensions(ws: Any, sheet: dict[str, Any]) -> None:
    column_widths = sheet.get("column_widths")
    if isinstance(column_widths, dict):
        for key, value in column_widths.items():
            if not isinstance(value, (int, float)):
                continue
            for column in expand_column_key(str(key)):
                ws.column_dimensions[column].width = value
    row_heights = sheet.get("row_heights")
    if isinstance(row_heights, dict):
        for key, value in row_heights.items():
            if isinstance(value, (int, float)) and str(key).isdigit():
                ws.row_dimensions[int(key)].height = value


def expand_column_key(key: str) -> list[str]:
    cleaned = key.strip().upper()
    if ":" not in cleaned:
        return [cleaned]
    start, end = cleaned.split(":", 1)
    start_index = column_index_from_string(start)
    end_index = column_index_from_string(end)
    if end_index < start_index:
        start_index, end_index = end_index, start_index
    return [get_column_letter(index) for index in range(start_index, end_index + 1)]


def apply_tables(ws: Any, tables: Any, used_table_names: set[str]) -> int:
    if not isinstance(tables, list):
        return 0
    count = 0
    for table_spec in tables:
        if not isinstance(table_spec, dict):
            continue
        name = safe_table_name(table_spec.get("name"), used_table_names)
        ref = str(table_spec.get("range") or "").strip()
        if not ref:
            raise ToolError("INVALID_ARGUMENT", "tables 每项都必须提供 range。")
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name=str(table_spec.get("style") or "TableStyleMedium9"),
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=read_bool(table_spec.get("show_row_stripes"), True),
            showColumnStripes=read_bool(table_spec.get("show_column_stripes"), False),
        )
        ws.add_table(table)
        count += 1
    return count


def has_table_specs(tables: Any) -> bool:
    return isinstance(tables, list) and any(isinstance(item, dict) for item in tables)


def apply_data_validations(ws: Any, validations: Any) -> int:
    if not isinstance(validations, list):
        return 0
    count = 0
    for spec in validations:
        if not isinstance(spec, dict):
            continue
        target_range = str(spec.get("range") or "").strip()
        validation_type = str(spec.get("type") or "").strip()
        if not target_range or not validation_type:
            raise ToolError("INVALID_ARGUMENT", "data_validations 每项都必须提供 range 和 type。")
        validation = DataValidation(
            type=validation_type,
            operator=spec.get("operator"),
            formula1=spec.get("formula1"),
            formula2=spec.get("formula2"),
            allow_blank=read_bool(spec.get("allow_blank"), True),
        )
        if isinstance(spec.get("prompt"), str):
            validation.prompt = spec["prompt"]
            validation.showInputMessage = True
        if isinstance(spec.get("error"), str):
            validation.error = spec["error"]
            validation.showErrorMessage = True
        validation.add(target_range)
        ws.add_data_validation(validation)
        count += 1
    return count


def apply_ranges(ws: Any, ranges: Any) -> int:
    if not isinstance(ranges, list):
        return 0
    count = 0
    for spec in ranges:
        if not isinstance(spec, dict):
            continue
        target_range = str(spec.get("range") or "").strip()
        if not target_range:
            raise ToolError("INVALID_ARGUMENT", "ranges 每项都必须提供 range。")
        count += apply_style_to_range(
            ws,
            target_range,
            spec.get("style"),
            spec.get("number_format") if isinstance(spec.get("number_format"), str) else None,
        )
    return count


def apply_page_setup(ws: Any, page: Any) -> None:
    if not isinstance(page, dict):
        return
    if isinstance(page.get("orientation"), str):
        ws.page_setup.orientation = page["orientation"]
    if isinstance(page.get("paper_size"), int):
        ws.page_setup.paperSize = page["paper_size"]
    if isinstance(page.get("fit_to_width"), int):
        ws.page_setup.fitToWidth = page["fit_to_width"]
    if isinstance(page.get("fit_to_height"), int):
        ws.page_setup.fitToHeight = page["fit_to_height"]
    if page.get("fit_to_width") is not None or page.get("fit_to_height") is not None:
        ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_sheet(
    wb: Workbook,
    sheet: dict[str, Any],
    used_table_names: set[str],
    warnings: list[str],
) -> dict[str, int | str]:
    ws = wb.create_sheet(validate_sheet_name(sheet.get("name"), set(wb.sheetnames)))
    tables = sheet.get("tables")
    tab_color = normalize_color(sheet.get("tab_color"))
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    if isinstance(sheet.get("freeze_panes"), str):
        ws.freeze_panes = sheet["freeze_panes"]
    if isinstance(sheet.get("auto_filter"), str):
        if has_table_specs(tables):
            warnings.append(f"{ws.title}: 已忽略 auto_filter；该 sheet 已定义 tables，表格会自带筛选。")
        else:
            ws.auto_filter.ref = sheet["auto_filter"]
    apply_dimensions(ws, sheet)

    cell_count = 0
    for block in sheet.get("data_blocks") or []:
        if isinstance(block, dict):
            cell_count += write_data_block(ws, block)
    for cell_spec in sheet.get("cells") or []:
        if isinstance(cell_spec, dict):
            cell_count += write_cell_spec(ws, cell_spec)
    for merge_range in sheet.get("merges") or []:
        if isinstance(merge_range, str):
            ws.merge_cells(merge_range)

    styled_cell_count = apply_ranges(ws, sheet.get("ranges"))
    table_count = apply_tables(ws, tables, used_table_names)
    validation_count = apply_data_validations(ws, sheet.get("data_validations"))

    conditional_count = 0
    for spec in sheet.get("conditional_formats") or []:
        if isinstance(spec, dict):
            add_conditional_format(ws, spec)
            conditional_count += 1

    chart_count = 0
    for spec in sheet.get("charts") or []:
        if isinstance(spec, dict):
            add_chart(ws, spec)
            chart_count += 1

    apply_page_setup(ws, sheet.get("page"))
    return {
        "name": ws.title,
        "cells_written": cell_count,
        "styled_cells": styled_cell_count,
        "tables": table_count,
        "data_validations": validation_count,
        "conditional_formats": conditional_count,
        "charts": chart_count,
    }


def safe_table_name(value: Any, used: set[str]) -> str:
    raw = str(value or "Table").strip()
    cleaned = re.sub(r"\W+", "_", raw)
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"Table_{cleaned}"
    candidate = cleaned[:240]
    base = candidate
    index = 2
    while candidate in used:
        candidate = f"{base}_{index}"
        index += 1
    used.add(candidate)
    return candidate


def _formula(value: str, location: str) -> str:
    return formula_value(value, location)


def validate_saved_workbook(path: Path) -> None:
    try:
        workbook = load_workbook(path, data_only=False, read_only=True)
        workbook.close()
    except Exception as exc:
        raise ToolError(
            "OUTPUT_INVALID",
            "生成后的 Excel 文件无法重新读取，已阻止输出坏文件。",
            {"output_path": str(path), "reason": str(exc) or type(exc).__name__},
        ) from exc


def create_workbook(payload: dict[str, Any], *, validate_only: bool = False) -> dict[str, Any]:
    payload = normalize_create_payload(payload)
    root = workspace_root()
    output_path = resolve_output_path(payload, root)
    overwrite = read_bool(payload.get("overwrite"), False)
    create_parent_dirs = read_bool(payload.get("create_parent_dirs"), True)
    if output_path.exists() and not overwrite:
        raise ToolError("OUTPUT_EXISTS", "输出文件已存在。", {"output_path": str(output_path)})
    if not validate_only:
        if create_parent_dirs:
            output_path.parent.mkdir(parents=True, exist_ok=True)
        elif not output_path.parent.is_dir():
            raise ToolError("DIRECTORY_NOT_FOUND", "输出目录不存在。", {"output_dir": str(output_path.parent)})
    elif not create_parent_dirs and not output_path.parent.is_dir():
        raise ToolError("DIRECTORY_NOT_FOUND", "输出目录不存在。", {"output_dir": str(output_path.parent)})

    sheets = payload.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise ToolError("INVALID_ARGUMENT", "sheets 必须是非空数组。")

    wb = Workbook()
    wb.remove(wb.active)
    apply_workbook_properties(wb, payload.get("properties"))
    used_table_names: set[str] = set()
    sheet_summaries: list[dict[str, int | str]] = []
    warnings: list[str] = []
    for sheet in sheets:
        if not isinstance(sheet, dict):
            raise ToolError("INVALID_ARGUMENT", "sheets 的每一项都必须是对象。")
        sheet_summaries.append(build_sheet(wb, sheet, used_table_names, warnings))

    active_sheet = payload.get("active_sheet")
    if isinstance(active_sheet, str) and active_sheet in wb.sheetnames:
        wb.active = wb.sheetnames.index(active_sheet)
    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True

    if not validate_only:
        wb.save(output_path)
        validate_saved_workbook(output_path)
    data = {
        "output_path": str(output_path),
        "sheet_count": len(wb.sheetnames),
        "sheets": sheet_summaries,
        "overwrite": overwrite,
        "validated_only": validate_only,
    }
    summary = "Excel 创建输入预检通过。" if validate_only else f"Excel 工作簿生成完成：{output_path.name}。"
    return ok(summary, data, warnings)


def normalize_create_payload(payload: dict[str, Any]) -> dict[str, Any]:
    if payload.get("mode") != "simple_table":
        return payload
    columns = payload.get("columns")
    rows = payload.get("rows")
    if not isinstance(columns, list) or not columns or not all(isinstance(item, str) for item in columns):
        raise ToolError("INVALID_ARGUMENT", "mode=simple_table 时 columns 必须是非空字符串数组。")
    if not isinstance(rows, list):
        raise ToolError("INVALID_ARGUMENT", "mode=simple_table 时 rows 必须是二维数组。")
    for row in rows:
        if not isinstance(row, list):
            raise ToolError("INVALID_ARGUMENT", "mode=simple_table 时 rows 的每一行都必须是数组。")
    sheet_name = payload.get("sheet_name") if isinstance(payload.get("sheet_name"), str) else "Sheet1"
    table_range = f"A1:{column_letter(len(columns))}{len(rows) + 1}"
    return {
        **payload,
        "sheets": [
            {
                "name": sheet_name.strip() or "Sheet1",
                "freeze_panes": "A2",
                "data_blocks": [
                    {
                        "start_cell": "A1",
                        "header": True,
                        "rows": [columns, *rows],
                        "header_style": {
                            "font": {"bold": True, "color": "FFFFFF"},
                            "fill": {"color": "4472C4"},
                            "alignment": {"horizontal": "center"},
                        },
                    }
                ],
                "tables": [
                    {
                        "name": "Table1",
                        "range": table_range,
                        "style": "TableStyleMedium9",
                    }
                ],
            }
        ],
    }


def column_letter(index: int) -> str:
    result = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result or "A"


def resolve_existing_workbook_path(payload: dict[str, Any], root: Path) -> Path:
    raw = payload.get("input_path")
    if not isinstance(raw, str) or not raw.strip():
        raise ToolError("INVALID_ARGUMENT", "input_path 必须是非空字符串。")
    path = Path(raw.strip()).expanduser()
    if not path.is_absolute():
        path = root / path
    resolved = path.resolve(strict=False)
    try:
        resolved.relative_to(root)
    except ValueError as exc:
        raise ToolError(
            "PATH_OUTSIDE_WORKSPACE",
            "input_path 必须位于工作区内。",
            {"input_path": str(resolved), "workspace_root": str(root)},
        ) from exc
    if not resolved.is_file():
        raise ToolError("INPUT_NOT_FOUND", "input_path 指向的 Excel 文件不存在。", {"input_path": str(resolved)})
    if resolved.suffix.lower() != ".xlsx":
        raise ToolError("INVALID_ARGUMENT", "目前只支持 .xlsx 文件。", {"input_path": str(resolved)})
    return resolved


def export_markdown_action(payload: dict[str, Any], root: Path) -> dict[str, Any]:
    input_path = resolve_workspace_path(payload.get("input_path"), root, label="input_path")
    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ToolError("INVALID_ARGUMENT", "input_path 必须是 .xlsx 或 .xlsm 文件。")
    if not input_path.is_file():
        raise ToolError("INPUT_NOT_FOUND", "input_path 指向的 Excel 文件不存在。", {"input_path": str(input_path)})
    content_path = resolve_workspace_path(payload.get("content_path"), root, label="content_path", suffix=".md")
    format_path = resolve_workspace_path(payload.get("format_path"), root, label="format_path", suffix=".md")
    assets_raw = payload.get("assets_dir")
    assets_dir = resolve_workspace_path(assets_raw, root, label="assets_dir") if assets_raw else content_path.parent / f"{content_path.stem}_assets"
    report_raw = payload.get("report_path")
    report_path = resolve_workspace_path(report_raw, root, label="report_path", suffix=".md") if report_raw else None
    overwrite = read_bool(payload.get("overwrite"), False)
    targets = [content_path, format_path] + ([report_path] if report_path else [])
    if not overwrite:
        existing = [str(path) for path in targets if path is not None and path.exists()]
        if existing:
            raise ToolError("OUTPUT_EXISTS", "反向提取目标文件已存在，请显式传 overwrite=true。", {"paths": existing})
    for path in targets:
        if path is not None:
            path.parent.mkdir(parents=True, exist_ok=True)
    selected = payload.get("sheets")
    selected_sheets = selected if isinstance(selected, list) and all(isinstance(item, str) for item in selected) else None
    return export_workbook(
        input_path,
        content_path,
        format_path,
        assets_dir,
        report_path,
        include_empty_rows=read_bool(payload.get("include_empty_rows"), False),
        selected_sheets=selected_sheets,
    )


def read_action(payload: dict[str, Any]) -> str:
    raw = payload.get("action")
    if not isinstance(raw, str) or not raw.strip():
        return "create"
    action = raw.strip().lower()
    aliases = {"read_summary": "inspect", "summary": "inspect", "build": "create"}
    return aliases.get(action, action)


def infer_validate_action(payload: dict[str, Any]) -> str:
    target = payload.get("target_action")
    if isinstance(target, str) and target.strip():
        return target.strip().lower()
    if isinstance(payload.get("operations"), list):
        return "edit"
    if isinstance(payload.get("sheets"), list):
        return "create"
    raise ToolError("INVALID_ARGUMENT", "validate 需要 target_action，或提供 sheets / operations 用于推断。")


def run(payload: dict[str, Any]) -> dict[str, Any]:
    try:
        root = workspace_root()
        action = read_action(payload)
        validate_only = read_bool(payload.get("validate_only"), False)

        if action == "inspect":
            data = inspect_workbook(resolve_existing_workbook_path(payload, root))
            return ok("Excel 工作簿摘要读取完成。", data)
        if action == "export_markdown":
            data = export_markdown_action(payload, root)
            return ok("Excel 已提取为双 Markdown，并附带 AI 读取规则。", data, data.get("warnings"))
        if action == "create":
            return create_workbook(payload, validate_only=validate_only)
        if action == "edit":
            data = edit_workbook(payload, root=root, validate_only=validate_only)
            summary = "Excel 编辑输入预检通过。" if validate_only else "Excel 工作簿编辑完成。"
            return ok(summary, data, data.get("warnings") if isinstance(data.get("warnings"), list) else [])
        if action == "validate":
            target_action = infer_validate_action(payload)
            if target_action == "create":
                return create_workbook(payload, validate_only=True)
            if target_action == "edit":
                data = edit_workbook(payload, root=root, validate_only=True)
                return ok("Excel 编辑输入预检通过。", data, data.get("warnings") if isinstance(data.get("warnings"), list) else [])
            raise ToolError("INVALID_ARGUMENT", "validate.target_action 只支持 create 或 edit。")
        raise ToolError(
            "INVALID_ACTION",
            "action 只支持 inspect、export_markdown、create、edit、validate。",
            {"action": action},
        )
    except ToolError as exc:
        return fail(exc.code, exc.message, exc.details)
    except Exception as exc:
        return fail("BUILD_FAILED", str(exc) or type(exc).__name__)


if __name__ == "__main__":
    for stream in (sys.stdin, sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8")
    run_tool(run)
