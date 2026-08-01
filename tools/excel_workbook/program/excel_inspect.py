from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def inspect_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        return {
            "input_path": str(path),
            "sheet_count": len(workbook.sheetnames),
            "active_sheet": workbook.active.title if workbook.active else None,
            "properties": _workbook_properties(workbook),
            "sheets": [_inspect_sheet(sheet) for sheet in workbook.worksheets],
        }
    finally:
        workbook.close()


def _workbook_properties(workbook: Any) -> dict[str, Any]:
    properties = workbook.properties
    keys = (
        "title",
        "subject",
        "creator",
        "keywords",
        "description",
        "category",
        "company",
        "manager",
        "lastModifiedBy",
    )
    return {key: value for key in keys if (value := getattr(properties, key, None))}


def _inspect_sheet(sheet: Any) -> dict[str, Any]:
    return {
        "name": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "used_range": sheet.calculate_dimension(),
        "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
        "auto_filter": sheet.auto_filter.ref,
        "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
        "tables": [
            {"name": table.name, "range": table.ref}
            for table in sheet.tables.values()
        ],
        "formula_count": _count_formulas(sheet),
        "chart_count": len(getattr(sheet, "_charts", []) or []),
        "conditional_format_count": len(sheet.conditional_formatting),
        "data_validation_count": len(sheet.data_validations.dataValidation),
    }


def _count_formulas(sheet: Any) -> int:
    count = 0
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                count += 1
    return count
