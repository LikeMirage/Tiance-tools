from __future__ import annotations

import copy
import json
import math
import os
import re
import sys
import tempfile
import urllib.request
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference, ScatterChart, Series
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from markdown_tables import MarkdownTableError, ParsedTable, extract_hidden_format, parse_tables, strip_hidden_format


class ToolError(ValueError):
    def __init__(self, code: str, message: str, details: dict[str, Any] | None = None):
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}


@dataclass
class FormatCell:
    merge: str | None = None
    props: dict[str, str] = field(default_factory=dict)


@dataclass
class SheetPlan:
    name: str
    data: ParsedTable
    fmt: ParsedTable | None
    cells: list[list[FormatCell]]
    merges: list[tuple[int, int, int, int]] = field(default_factory=list)
    styles: dict[tuple[int, int], dict[str, str]] = field(default_factory=dict)
    range_styles: list[tuple[str, dict[str, str]]] = field(default_factory=list)
    images: list[tuple[int, int, dict[str, str]]] = field(default_factory=list)
    widths: dict[int, str] = field(default_factory=dict)
    heights: dict[int, str] = field(default_factory=dict)
    sheet_props: dict[str, list[tuple[str, int, int]]] = field(default_factory=dict)
    validations: list[tuple[str, str]] = field(default_factory=list)
    conditionals: list[tuple[str, str, str, str | None]] = field(default_factory=list)
    charts: list[tuple[int, int, dict[str, str]]] = field(default_factory=list)

    @property
    def rows(self) -> int:
        return len(self.data.rows)

    @property
    def cols(self) -> int:
        return len(self.data.rows[0])


_SHEET_BAD = re.compile(r"[\\/*?:\[\]]")
_CELL_REF = re.compile(r"^[A-Za-z]{1,3}[1-9]\d*$")
_RANGE_REF = re.compile(r"^[A-Za-z]{1,3}[1-9]\d*(?::[A-Za-z]{1,3}[1-9]\d*)?$")
_HEX = re.compile(r"^#(?:[0-9a-f]{3}|[0-9a-f]{6}|[0-9a-f]{8})$", re.I)
_NUMBER = re.compile(r"^[+-]?(?:0|[1-9]\d*)(?:\.\d+)?$")
_DIRECTIONS = {"上", "下", "左", "右", "上直到", "下直到", "左直到", "右直到"}
_ALIASES = {
    "bold": ("font-weight", "bold"),
    "italic": ("font-style", "italic"),
    "underline": ("text-decoration", "underline"),
    "wrap": ("wrap-text", "on"),
    "nowrap": ("wrap-text", "off"),
}
_KNOWN_KEYS = {
    "font-weight", "font-style", "text-decoration", "font-size", "font-family", "font-color", "color",
    "fill", "background", "text-align", "vertical-align", "wrap-text", "white-space", "shrink-to-fit",
    "text-rotation", "border", "border-top", "border-right", "border-bottom", "border-left", "border-color",
    "number-format", "format", "protection", "width", "height", "column-width", "row-height", "freeze",
    "filter", "table", "table-style", "tab-color", "hidden-row", "hidden-column", "print-area",
    "page-orientation", "fit-to-width", "fit-to-height", "margin", "merge", "image", "image-width", "image-height",
    "hyperlink", "comment", "validation", "data-bar", "color-scale", "cell-is", "formula-rule", "chart",
    "chart-title", "chart-width", "chart-height", "chart-cats", "chart-series", "chart-x", "chart-y", "range",
}


def workspace_root() -> Path:
    raw = os.environ.get("TIANCE_WORKSPACE_ROOT") or os.environ.get("WORKSPACE_ROOT") or os.getcwd()
    return Path(raw).expanduser().resolve()


def safe_path(raw: Any, root: Path, suffix: str) -> Path:
    if not isinstance(raw, str) or not raw.strip():
        raise ToolError("INVALID_ARGUMENT", "路径必须是非空字符串。")
    path = Path(raw.strip()).expanduser()
    if not path.is_absolute():
        path = root / path
    result = path.resolve(strict=False)
    try:
        result.relative_to(root)
    except ValueError as exc:
        raise ToolError("PATH_OUTSIDE_WORKSPACE", "路径必须位于工作区内。") from exc
    if result.suffix.lower() != suffix:
        result = result.with_suffix(suffix)
    return result


def color(value: str) -> str:
    value = value.strip()
    if not _HEX.fullmatch(value):
        raise ToolError("INVALID_FORMAT", f"颜色必须是 #RGB、#RRGGBB 或 #RRGGBBAA：{value}")
    raw = value[1:].upper()
    if len(raw) == 3:
        raw = "".join(char * 2 for char in raw)
    return raw


def parse_directive(text: str, where: str) -> FormatCell:
    result = FormatCell()
    if not text.strip():
        return result
    if any(value in text for value in ("；", "\n", "\r")):
        raise ToolError("INVALID_FORMAT", f"{where}: 指令只能使用英文分号且不能换行。")
    for raw in text.split(";"):
        token = raw.strip()
        if not token:
            continue
        if ":" not in token:
            alias = _ALIASES.get(token.casefold())
            if alias is None:
                raise ToolError("INVALID_FORMAT", f"{where}: 未知单 token 指令 {token!r}，合并必须写 merge:名称。")
            key, value = alias
        else:
            key, value = (part.strip() for part in token.split(":", 1))
            key = key.casefold()
            if key not in _KNOWN_KEYS:
                raise ToolError("INVALID_FORMAT", f"{where}: 未知格式指令 {key!r}。")
            if not value:
                raise ToolError("INVALID_FORMAT", f"{where}: {key} 缺少值。")
        if key == "merge":
            if result.merge is not None:
                raise ToolError("INVALID_FORMAT", f"{where}: 同一格不能有多个 merge 指令。")
            result.merge = value
            continue
        if key in result.props:
            raise ToolError("INVALID_FORMAT", f"{where}: 指令 {key} 重复。")
        _validate_value(key, value, where)
        result.props[key] = value
    return result


def _validate_value(key: str, value: str, where: str) -> None:
    allowed_pipe = key in {"cell-is", "formula-rule", "validation"}
    forbidden = ("\n", "\r") if key in {"cell-is", "formula-rule"} else ("\n", "\r", "<", ">")
    if any(char in value for char in forbidden) or ("|" in value and not allowed_pipe):
        raise ToolError("INVALID_FORMAT", f"{where}: {key} 的值包含禁止字符。")
    if key in {"color", "font-color", "background", "fill", "border-color", "tab-color"}:
        color(value)
    elif key in {"font-weight"} and value.casefold() not in {"bold", "normal"}:
        raise ToolError("INVALID_FORMAT", f"{where}: font-weight 只能是 bold 或 normal。")
    elif key in {"font-style"} and value.casefold() not in {"italic", "normal"}:
        raise ToolError("INVALID_FORMAT", f"{where}: font-style 只能是 italic 或 normal。")
    elif key == "text-decoration" and value.casefold() not in {"underline", "none"}:
        raise ToolError("INVALID_FORMAT", f"{where}: text-decoration 只能是 underline 或 none。")
    elif key in {"text-align"} and value.casefold() not in {"left", "center", "right", "justify"}:
        raise ToolError("INVALID_FORMAT", f"{where}: 水平对齐值无效。")
    elif key == "vertical-align" and value.casefold() not in {"top", "center", "middle", "bottom"}:
        raise ToolError("INVALID_FORMAT", f"{where}: 垂直对齐值无效。")
    elif key in {"wrap-text", "white-space", "shrink-to-fit"} and value.casefold() not in {"on", "off", "normal", "nowrap", "true", "false"}:
        raise ToolError("INVALID_FORMAT", f"{where}: {key} 的值无效。")
    elif key.startswith("border") and key != "border-color" and value.casefold() not in {"thin", "medium", "thick", "double", "dashed", "dotted", "none"}:
        raise ToolError("INVALID_FORMAT", f"{where}: 边框样式无效。")
    elif key in {"font-size", "image-width", "image-height", "width", "height", "column-width", "row-height", "chart-width", "chart-height", "fit-to-width", "fit-to-height"}:
        if value.casefold() not in {"auto", "none"}:
            try:
                if float(value.removesuffix("px")) <= 0:
                    raise ValueError
            except ValueError as exc:
                raise ToolError("INVALID_FORMAT", f"{where}: {key} 必须是正数或 auto。") from exc
    elif key in {"freeze", "filter", "print-area", "range", "hyperlink", "image", "chart", "validation", "data-bar", "color-scale", "cell-is", "formula-rule", "comment", "table-style", "font-family", "number-format", "format", "protection", "margin", "page-orientation", "chart-title", "chart-cats", "chart-series", "chart-x", "chart-y"}:
        return


def parse_format_table(table: ParsedTable) -> list[list[FormatCell]]:
    return [
        [parse_directive(value, f"{table.title or 'Sheet'}!{get_column_letter(col)}{row}") for col, value in enumerate(line, 1)]
        for row, line in enumerate(table.rows, 1)
    ]


def sheet_name(table: ParsedTable, index: int) -> str:
    name = (table.title or f"Sheet{index}").strip()
    if not name or len(name) > 31 or _SHEET_BAD.search(name):
        raise ToolError("INVALID_SHEET_NAME", f"第 {index} 个表的 Sheet 名非法：{name!r}")
    return name


def match_format_tables(data_tables: list[ParsedTable], format_tables: list[ParsedTable]) -> dict[int, ParsedTable]:
    result: dict[int, ParsedTable] = {}
    by_title: dict[str, int] = {}
    for index, table in enumerate(data_tables):
        title = (table.title or f"Sheet{index + 1}").casefold()
        if title in by_title:
            raise ToolError("DUPLICATE_SHEET_NAME", f"内容文件存在重复表格标题：{table.title or title}")
        by_title[title] = index
    used: set[int] = set()
    unnamed_index = 0
    for table in format_tables:
        if table.title:
            index = by_title.get(table.title.casefold())
            if index is None:
                raise ToolError("FORMAT_TABLE_NOT_FOUND", f"格式文件中的表格 {table.title!r} 在内容文件中不存在。")
        else:
            while unnamed_index in used:
                unnamed_index += 1
            if unnamed_index >= len(data_tables):
                raise ToolError("FORMAT_TABLE_NOT_FOUND", "格式文件中的无标题表格没有对应内容表。")
            index = unnamed_index
            unnamed_index += 1
        if index in used:
            raise ToolError("DUPLICATE_FORMAT_TABLE", f"格式文件重复覆盖表格 {data_tables[index].title or index + 1}。")
        used.add(index)
        result[index] = table
    return result


def build_plans(data_tables: list[ParsedTable], format_tables: list[ParsedTable]) -> list[SheetPlan]:
    matches = match_format_tables(data_tables, format_tables)
    names: set[str] = set()
    plans: list[SheetPlan] = []
    for index, data in enumerate(data_tables):
        name = sheet_name(data, index + 1)
        if name.casefold() in names:
            raise ToolError("DUPLICATE_SHEET_NAME", f"重复 Sheet 名：{name}")
        names.add(name.casefold())
        fmt = matches.get(index)
        if fmt is not None:
            if len(fmt.rows) != len(data.rows) or len(fmt.rows[0]) != len(data.rows[0]):
                raise ToolError(
                    "TABLE_SHAPE_MISMATCH",
                    f"表格 {name} 的内容和格式必须同构：内容 {len(data.rows)}x{len(data.rows[0])}，格式 {len(fmt.rows)}x{len(fmt.rows[0])}。",
                )
            cells = parse_format_table(fmt)
        else:
            cells = [[FormatCell() for _ in data.rows[0]] for _ in data.rows]
        plan = SheetPlan(name=name, data=data, fmt=fmt, cells=cells)
        collect_plan_features(plan)
        plans.append(plan)
    return plans


def collect_plan_features(plan: SheetPlan) -> None:
    markers: dict[str, list[tuple[int, int]]] = {}
    occupied: dict[tuple[int, int], str] = {}
    for row, line in enumerate(plan.cells, 1):
        for col, cell in enumerate(line, 1):
            if cell.merge:
                markers.setdefault(cell.merge.casefold(), []).append((row, col))
            props = dict(cell.props)
            range_ref_value = props.pop("range", None)
            if range_ref_value:
                _validate_range(range_ref_value, plan.name)
                plan.range_styles.append((range_ref_value.upper(), _style_only(props)))
            direct_style = _style_only(props)
            if direct_style:
                plan.styles[(row, col)] = direct_style
            if "image" in props:
                plan.images.append((row, col, props))
            if "chart" in props:
                plan.charts.append((row, col, props))
            for key in ("width", "column-width"):
                if key in props:
                    _collect_dimension(plan.widths, col, props[key], f"{plan.name}!{get_column_letter(col)}{row}")
            for key in ("height", "row-height"):
                if key in props:
                    _collect_dimension(plan.heights, row, props[key], f"{plan.name}!{get_column_letter(col)}{row}")
            for key in ("freeze", "filter", "table", "table-style", "tab-color", "hidden-row", "hidden-column", "print-area", "page-orientation", "fit-to-width", "fit-to-height", "margin"):
                if key in props:
                    plan.sheet_props.setdefault(key, []).append((props[key], row, col))
            if "validation" in props:
                raw = props["validation"]
                if "@" not in raw:
                    raise ToolError("INVALID_FORMAT", f"{plan.name}!{get_column_letter(col)}{row}: validation 必须写成 选项1,选项2@A2:A10。")
                choices, target = raw.rsplit("@", 1)
                _validate_range(target, plan.name)
                plan.validations.append((choices.strip(), target.upper()))
            for key in ("data-bar", "color-scale", "cell-is", "formula-rule"):
                if key in props:
                    raw = props[key]
                    if "@" not in raw:
                        raise ToolError("INVALID_FORMAT", f"{plan.name}!{get_column_letter(col)}{row}: {key} 必须包含 @范围。")
                    rule, target = raw.rsplit("@", 1)
                    _validate_range(target, plan.name)
                    plan.conditionals.append((key, rule, target.upper(), None))
    for marker, coords in markers.items():
        if len(coords) == 1:
            continue
        rect = (min(row for row, _ in coords), min(col for _, col in coords), max(row for row, _ in coords), max(col for _, col in coords))
        for row in range(rect[0], rect[2] + 1):
            for col in range(rect[1], rect[3] + 1):
                current = plan.cells[row - 1][col - 1].merge
                if current and current.casefold() != marker:
                    raise ToolError("MERGE_CROSSING", f"{plan.name}: 合并 {marker!r} 的包围盒内出现 {current!r}。")
                previous = occupied.get((row, col))
                if previous is not None and previous != marker:
                    raise ToolError("MERGE_OVERLAP", f"{plan.name}: 合并 {marker!r} 与 {previous!r} 重叠。")
                occupied[(row, col)] = marker
                if (row, col) != (rect[0], rect[1]) and parse_value(plan.data.rows[row - 1][col - 1]) not in (None, ""):
                    raise ToolError("MERGE_DATA_LOSS", f"{plan.name}!{get_column_letter(col)}{row}: 合并区覆盖了非空内容。请只保留左上角内容。")
        plan.merges.append(rect)


def _style_only(props: dict[str, str]) -> dict[str, str]:
    ignored = {"image", "image-width", "image-height", "chart", "chart-title", "chart-width", "chart-height", "chart-cats", "chart-series", "chart-x", "chart-y"}
    return {key: value for key, value in props.items() if key not in ignored and key not in {"range"}}


def _collect_dimension(values: dict[int, str], key: int, value: str, where: str) -> None:
    if key in values and values[key].casefold() != value.casefold():
        raise ToolError("CONFLICTING_DIMENSION", f"{where}: 同一行或列出现冲突尺寸 {values[key]!r} 与 {value!r}。")
    values[key] = value


def _validate_range(value: str, sheet: str) -> None:
    if not _RANGE_REF.fullmatch(value.strip()):
        raise ToolError("INVALID_RANGE", f"{sheet}: 非法范围 {value!r}。")


def parse_value(value: str) -> Any:
    value = value.strip().replace("<br>", "\n")
    if not value:
        return None
    if value.lower().startswith("text:"):
        return value[5:]
    if value.startswith("'"):
        return value[1:]
    if value.lower().startswith("date:"):
        try:
            return datetime.strptime(value[5:], "%Y-%m-%d").date()
        except ValueError as exc:
            raise ToolError("INVALID_VALUE", f"非法日期：{value}") from exc
    if value.lower() in {"true", "false"}:
        return value.lower() == "true"
    if value.startswith("="):
        return value
    if _NUMBER.fullmatch(value):
        integer = value.lstrip("+-").split(".", 1)[0]
        if len(integer) > 1 and integer.startswith("0"):
            return value
        return float(value) if "." in value else int(value)
    return value


def apply_style(cell: Any, props: dict[str, str]) -> None:
    current_font = copy.copy(cell.font)
    font_color = color(props.get("font-color", props.get("color", "#000000"))) if props.get("font-color") or props.get("color") else current_font.color
    if any(key in props for key in ("font-weight", "font-style", "text-decoration", "font-size", "font-family", "font-color", "color")):
        cell.font = Font(
            name=props.get("font-family", current_font.name),
            sz=float(str(props.get("font-size", current_font.sz or 11)).removesuffix("px")),
            bold=props.get("font-weight", "bold" if current_font.bold else "normal").casefold() == "bold",
            italic=props.get("font-style", "italic" if current_font.italic else "normal").casefold() == "italic",
            underline="single" if props.get("text-decoration", current_font.underline or "none").casefold() == "underline" else None,
            strike=current_font.strike,
            color=font_color,
        )
    fill_value = props.get("fill", props.get("background"))
    if fill_value:
        cell.fill = PatternFill(fill_type="solid", fgColor=color(fill_value))
    if any(key in props for key in ("text-align", "vertical-align", "wrap-text", "white-space", "shrink-to-fit", "text-rotation")):
        alignment = copy.copy(cell.alignment)
        alignment.horizontal = props.get("text-align", alignment.horizontal)
        alignment.vertical = {"middle": "center"}.get(props.get("vertical-align", alignment.vertical), props.get("vertical-align", alignment.vertical))
        if "wrap-text" in props:
            alignment.wrap_text = props["wrap-text"].casefold() in {"on", "true"}
        elif "white-space" in props:
            alignment.wrap_text = props["white-space"].casefold() == "normal"
        if "shrink-to-fit" in props:
            alignment.shrink_to_fit = props["shrink-to-fit"].casefold() in {"on", "true"}
        if "text-rotation" in props:
            alignment.text_rotation = int(props["text-rotation"])
        cell.alignment = alignment
    border_color = color(props.get("border-color", "#000000"))
    border = copy.copy(cell.border)
    sides = {}
    if "border" in props:
        side = _make_side(props["border"], border_color)
        sides = {"left": side, "right": side, "top": side, "bottom": side}
    for key, name in (("border-left", "left"), ("border-right", "right"), ("border-top", "top"), ("border-bottom", "bottom")):
        if key in props:
            sides[name] = _make_side(props[key], border_color)
    if sides:
        cell.border = Border(
            left=sides.get("left", border.left), right=sides.get("right", border.right),
            top=sides.get("top", border.top), bottom=sides.get("bottom", border.bottom),
            diagonal=border.diagonal, diagonal_direction=border.diagonal_direction,
        )
    if props.get("number-format") or props.get("format"):
        cell.number_format = props.get("number-format", props.get("format")) or "General"
    if "protection" in props:
        cell.protection = Protection(locked=props["protection"].casefold() != "unlocked")
    if "hyperlink" in props:
        cell.hyperlink = props["hyperlink"]
    if "comment" in props:
        from openpyxl.comments import Comment
        cell.comment = Comment(props["comment"], "Tiance")


def _make_side(style: str, colour: str) -> Side:
    return Side(style=None if style.casefold() == "none" else style.casefold(), color=colour)


def merged_owner(plan: SheetPlan, row: int, col: int) -> tuple[int, int]:
    for top, left, bottom, right in plan.merges:
        if top <= row <= bottom and left <= col <= right:
            return top, left
    return row, col


def apply_dimensions(ws: Any, plan: SheetPlan) -> None:
    for col, value in plan.widths.items():
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = auto_width(ws, col) if value.casefold() == "auto" else float(value)
    for row, value in plan.heights.items():
        ws.row_dimensions[row].height = auto_height(ws, row) if value.casefold() == "auto" else float(value)


def auto_width(ws: Any, col: int) -> float:
    maximum = 8.43
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, col).value
        if value is None:
            continue
        longest = max((sum(2 if ord(char) > 255 else 1 for char in line) for line in str(value).splitlines()), default=0)
        maximum = max(maximum, longest * float(ws.cell(row, col).font.sz or 11) / 11 + 2)
    return min(maximum, 80)


def auto_height(ws: Any, row: int) -> float:
    height = 15.0
    for cell in ws[row]:
        if cell.value is None:
            continue
        lines = str(cell.value).count("\n") + 1
        if cell.alignment.wrap_text:
            width = ws.column_dimensions[get_column_letter(cell.column)].width or 8.43
            visual = max((sum(2 if ord(char) > 255 else 1 for char in line) for line in str(cell.value).splitlines()), default=1)
            lines = max(lines, math.ceil(visual / max(width - 1, 1)))
        height = max(height, lines * float(cell.font.sz or 11) * 1.35)
    return min(max(height, 15), 300)


def apply_sheet_properties(ws: Any, plan: SheetPlan) -> None:
    props = plan.sheet_props
    if "freeze" in props:
        ws.freeze_panes = _one_value(props["freeze"], "freeze", plan.name)
    if "filter" in props:
        ws.auto_filter.ref = _one_value(props["filter"], "filter", plan.name)
    if "tab-color" in props:
        ws.sheet_properties.tabColor = color(_one_value(props["tab-color"], "tab-color", plan.name))
    if "print-area" in props:
        ws.print_area = _one_value(props["print-area"], "print-area", plan.name)
    if "page-orientation" in props:
        ws.page_setup.orientation = _one_value(props["page-orientation"], "page-orientation", plan.name)
    if "fit-to-width" in props:
        ws.page_setup.fitToWidth = int(_one_value(props["fit-to-width"], "fit-to-width", plan.name))
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    if "fit-to-height" in props:
        ws.page_setup.fitToHeight = int(_one_value(props["fit-to-height"], "fit-to-height", plan.name))
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    if "margin" in props:
        margin = float(_one_value(props["margin"], "margin", plan.name))
        ws.page_margins.left = ws.page_margins.right = margin
        ws.page_margins.top = ws.page_margins.bottom = margin
    for value, row, _ in props.get("hidden-row", []):
        ws.row_dimensions[row].hidden = value.casefold() not in {"off", "false", "0"}
    for value, _, col in props.get("hidden-column", []):
        ws.column_dimensions[get_column_letter(col)].hidden = value.casefold() not in {"off", "false", "0"}


def _one_value(values: list[tuple[str, int, int]], key: str, sheet: str) -> str:
    unique = {value.casefold(): value for value, _, _ in values}
    if len(unique) != 1:
        raise ToolError("CONFLICTING_SHEET_PROPERTY", f"{sheet}: {key} 存在冲突值。")
    return next(iter(unique.values()))


def apply_table_object(ws: Any, plan: SheetPlan) -> None:
    values = plan.sheet_props.get("table", [])
    if not values:
        return
    if _one_value(values, "table", plan.name).casefold() != "on":
        return
    if plan.merges and any(top == 1 for top, _, _, _ in plan.merges):
        raise ToolError("TABLE_HEADER_MERGED", f"{plan.name}: table:on 时第一行不能合并。")
    headers = [str(plan.data.rows[0][col]).strip() for col in range(plan.cols)]
    if any(not header for header in headers) or len({header.casefold() for header in headers}) != len(headers):
        raise ToolError("TABLE_HEADER_INVALID", f"{plan.name}: table:on 的表头必须非空且不重复。")
    display_name = re.sub(r"[^A-Za-z0-9_]", "", f"Table_{plan.name}")[:200] or "Table1"
    existing = {
        table.name.casefold()
        for sheet in ws.parent.worksheets
        for table in sheet.tables.values()
    }
    base_name = display_name
    counter = 2
    while display_name.casefold() in existing:
        suffix = f"_{counter}"
        display_name = f"{base_name[:200 - len(suffix)]}{suffix}"
        counter += 1
    table = Table(displayName=display_name, ref=f"A1:{get_column_letter(plan.cols)}{plan.rows}")
    table.tableStyleInfo = TableStyleInfo(
        name=_one_value(plan.sheet_props.get("table-style", [("TableStyleMedium9", 1, 1)]), "table-style", plan.name),
        showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def apply_validations(ws: Any, plan: SheetPlan) -> None:
    for choices, target in plan.validations:
        escaped = choices.replace('"', '""')
        if len(escaped) > 255:
            raise ToolError("VALIDATION_TOO_LONG", f"{plan.name}: 下拉选项总长度超过 Excel 限制。")
        validation = DataValidation(type="list", formula1=f'"{escaped}"', allow_blank=True)
        validation.error = "请选择列表中的值"
        validation.errorTitle = "输入无效"
        validation.showErrorMessage = True
        ws.add_data_validation(validation)
        validation.add(target)


def apply_conditionals(ws: Any, plan: SheetPlan) -> None:
    for kind, raw, target, _ in plan.conditionals:
        if kind == "data-bar":
            ws.conditional_formatting.add(target, DataBarRule(start_type="min", end_type="max", color=color(raw.split("|", 1)[0])))
        elif kind == "color-scale":
            values = raw.split(",")
            if len(values) != 3:
                raise ToolError("INVALID_FORMAT", f"{plan.name}: color-scale 必须是三个颜色。")
            ws.conditional_formatting.add(target, ColorScaleRule(start_type="min", start_color=color(values[0]), mid_type="percentile", mid_value=50, mid_color=color(values[1]), end_type="max", end_color=color(values[2])))
        elif kind == "cell-is":
            condition, _, colour = raw.partition("|")
            match = re.match(r"^(>=|<=|<>|=|>|<)\s*(.+)$", condition.strip())
            if not match or not colour.strip():
                raise ToolError("INVALID_FORMAT", f"{plan.name}: cell-is 必须写成 运算符|颜色。")
            ws.conditional_formatting.add(
                target,
                CellIsRule(
                    operator=match.group(1),
                    formula=[match.group(2).strip()],
                    fill=PatternFill("solid", fgColor=color(colour.strip())),
                ),
            )
        elif kind == "formula-rule":
            formula, _, colour = raw.partition("|")
            ws.conditional_formatting.add(target, FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=color(colour or "#FFFF00"))))


def apply_charts(ws: Any, plan: SheetPlan) -> None:
    for row, col, props in plan.charts:
        chart_text = props["chart"]
        chart_type, _, raw_range = chart_text.partition("@")
        if not raw_range:
            raise ToolError("INVALID_FORMAT", f"{plan.name}: chart 必须写成 chart:bar@A1:D10。")
        min_col, min_row, max_col, max_row = range_boundaries(raw_range.upper())
        if chart_type.casefold() == "bar":
            chart = BarChart()
        elif chart_type.casefold() == "line":
            chart = LineChart()
        elif chart_type.casefold() == "area":
            chart = AreaChart()
        elif chart_type.casefold() == "pie":
            chart = PieChart()
        elif chart_type.casefold() == "scatter":
            chart = ScatterChart()
        else:
            raise ToolError("INVALID_FORMAT", f"{plan.name}: 不支持的图表类型 {chart_type!r}。")
        cats = props.get("chart-cats", "col1").casefold()
        if isinstance(chart, PieChart) or cats == "col1":
            data = Reference(ws, min_col=min_col + 1, min_row=min_row, max_col=max_col, max_row=max_row)
            categories = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
        else:
            data = Reference(ws, min_col=min_col, min_row=min_row + 1, max_col=max_col, max_row=max_row)
            categories = Reference(ws, min_col=min_col + 1, min_row=min_row, max_col=max_col)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = props.get("chart-title")
        chart.x_axis.title = props.get("chart-x")
        chart.y_axis.title = props.get("chart-y")
        if props.get("chart-width"):
            chart.width = float(props["chart-width"])
        if props.get("chart-height"):
            chart.height = float(props["chart-height"])
        if isinstance(chart, PieChart):
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
        ws.add_chart(chart, f"{get_column_letter(col)}{row}")


def resolve_image(source: str, format_path: Path, temporary: list[Path]) -> Path:
    if re.match(r"^https?://", source, re.I):
        suffix = Path(source.split("?", 1)[0]).suffix or ".img"
        handle, name = tempfile.mkstemp(suffix=suffix)
        os.close(handle)
        target = Path(name)
        temporary.append(target)
        request = urllib.request.Request(source, headers={"User-Agent": "Tiance-md-to-excel/2.0"})
        with urllib.request.urlopen(request, timeout=15) as response:
            content = response.read(10 * 1024 * 1024 + 1)
        if len(content) > 10 * 1024 * 1024:
            raise ToolError("IMAGE_TOO_LARGE", "网络图片超过 10MB。")
        target.write_bytes(content)
        return target
    path = Path(source)
    if not path.is_absolute():
        path = (format_path.parent / path).resolve()
    if not path.is_file():
        raise ToolError("IMAGE_NOT_FOUND", f"找不到图片：{path}")
    if path.stat().st_size > 10 * 1024 * 1024:
        raise ToolError("IMAGE_TOO_LARGE", f"图片超过 10MB：{path}")
    return path


def render_plan(ws: Any, plan: SheetPlan, format_path: Path | None, warnings: list[str]) -> None:
    values = [[parse_value(value) for value in row] for row in plan.data.rows]
    for row in range(1, plan.rows + 1):
        for col in range(1, plan.cols + 1):
            ws.cell(row, col, values[row - 1][col - 1])
    for top, left, bottom, right in plan.merges:
        ws.merge_cells(start_row=top, start_column=left, end_row=bottom, end_column=right)
    for (row, col), props in plan.styles.items():
        owner_row, owner_col = merged_owner(plan, row, col)
        apply_style(ws.cell(owner_row, owner_col), props)
    for ref, props in plan.range_styles:
        for line in ws[ref]:
            for cell in line:
                apply_style(cell, props)
    apply_dimensions(ws, plan)
    apply_sheet_properties(ws, plan)
    apply_table_object(ws, plan)
    apply_validations(ws, plan)
    apply_conditionals(ws, plan)
    apply_charts(ws, plan)
    if format_path is not None:
        temporary: list[Path] = []
        try:
            for row, col, props in plan.images:
                source = props["image"]
                if values[row - 1][col - 1] not in (None, ""):
                    raise ToolError(
                        "IMAGE_DATA_CONFLICT",
                        f"{plan.name}!{get_column_letter(col)}{row}: image 单元格不能同时有正文内容。",
                    )
                try:
                    image = XLImage(str(resolve_image(source, format_path, temporary)))
                    owner_row, owner_col = merged_owner(plan, row, col)
                    if props.get("image-width"):
                        image.width = float(props["image-width"].removesuffix("px"))
                    if props.get("image-height"):
                        image.height = float(props["image-height"].removesuffix("px"))
                    ws.add_image(image, f"{get_column_letter(owner_col)}{owner_row}")
                except ToolError:
                    raise
                except Exception as exc:
                    raise ToolError("IMAGE_INSERT_FAILED", f"{plan.name}!{get_column_letter(col)}{row}: 图片插入失败：{exc}") from exc
        finally:
            for path in temporary:
                path.unlink(missing_ok=True)


def convert(payload: dict[str, Any]) -> dict[str, Any]:
    root = workspace_root()
    content_path = safe_path(payload.get("content_path"), root, ".md")
    output_path = safe_path(payload.get("output_path"), root, ".xlsx")
    format_raw = payload.get("format_path")
    format_path = safe_path(format_raw, root, ".md") if format_raw else None
    if not content_path.is_file():
        raise ToolError("CONTENT_NOT_FOUND", f"找不到内容文件：{content_path}")
    if format_path is not None and not format_path.is_file():
        raise ToolError("FORMAT_NOT_FOUND", f"找不到格式文件：{format_path}")
    if output_path.exists() and not payload.get("overwrite", False):
        raise ToolError("OUTPUT_EXISTS", f"输出文件已存在：{output_path}，如需覆盖请显式传 overwrite=true。")
    try:
        content_text = content_path.read_text(encoding="utf-8-sig")
        format_text = format_path.read_text(encoding="utf-8-sig") if format_path else extract_hidden_format(content_text)
        data_tables = parse_tables(strip_hidden_format(content_text))
        format_tables = parse_tables(format_text) if format_text else []
    except UnicodeDecodeError as exc:
        raise ToolError("INPUT_INVALID", "Markdown 必须使用 UTF-8 编码。") from exc
    except MarkdownTableError as exc:
        raise ToolError("MARKDOWN_INVALID", str(exc)) from exc
    if not data_tables:
        raise ToolError("NO_TABLES_FOUND", "内容 Markdown 中没有找到标准表格。")
    plans = build_plans(data_tables, format_tables)
    workbook = Workbook()
    workbook.remove(workbook.active)
    warnings: list[str] = []
    for plan in plans:
        render_plan(workbook.create_sheet(plan.name), plan, format_path or content_path, warnings)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(prefix=f".{output_path.stem}-", suffix=".xlsx", dir=output_path.parent, delete=False) as handle:
            temporary_path = Path(handle.name)
        workbook.save(temporary_path)
        checked = load_workbook(temporary_path, data_only=False)
        if [sheet.title for sheet in checked.worksheets] != [plan.name for plan in plans]:
            raise ToolError("OUTPUT_VERIFY_FAILED", "生成文件的 Sheet 顺序与内容文件不一致。")
        checked.close()
        os.replace(temporary_path, output_path)
        temporary_path = None
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)
    return {
        "output_path": str(output_path),
        "content_path": str(content_path),
        "format_path": str(format_path) if format_path else None,
        "table_count": len(plans),
        "warnings": warnings,
        "sheets": [{"name": plan.name, "rows": plan.rows, "columns": plan.cols, "merges": len(plan.merges)} for plan in plans],
    }


def main() -> None:
    try:
        payload = json.load(sys.stdin)
        if not isinstance(payload, dict):
            raise ToolError("INVALID_ARGUMENT", "输入必须是 JSON 对象。")
        print(json.dumps({"ok": True, "summary": "Markdown 已转换为 Excel。", "data": convert(payload)}, ensure_ascii=False))
    except ToolError as exc:
        print(json.dumps({"ok": False, "error": f"{exc.code}: {exc.message}", "error_info": {"code": exc.code, "message": exc.message, "details": exc.details}}, ensure_ascii=False))
    except (MarkdownTableError, ValueError) as exc:
        print(json.dumps({"ok": False, "error": f"INVALID_FORMAT: {exc}", "error_info": {"code": "INVALID_FORMAT", "message": str(exc), "details": {}}}, ensure_ascii=False))
    except Exception as exc:
        print(json.dumps({"ok": False, "error": f"INTERNAL_ERROR: {exc}", "error_info": {"code": "INTERNAL_ERROR", "message": str(exc), "details": {}}}, ensure_ascii=False))


if __name__ == "__main__":
    main()
