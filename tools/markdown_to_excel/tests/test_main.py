from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).parents[4]
PROGRAM = Path(__file__).parent.parent / "program" / "main.py"
CONTENT = Path(__file__).parent.parent / "examples" / "complete-content.md"
FORMAT = Path(__file__).parent.parent / "examples" / "complete-format.md"


def run_tool(payload: dict, root: Path = ROOT) -> dict:
    result = subprocess.run(
        [sys.executable, str(PROGRAM)],
        input=json.dumps(payload),
        text=True,
        capture_output=True,
        env={**os.environ, "TIANCE_WORKSPACE_ROOT": str(root)},
        check=False,
    )
    return json.loads(result.stdout)


def test_complete_example_round_trip(tmp_path):
    content = tmp_path / "content.md"
    format_file = tmp_path / "format.md"
    content.write_text(CONTENT.read_text(encoding="utf-8"), encoding="utf-8")
    format_file.write_text(FORMAT.read_text(encoding="utf-8"), encoding="utf-8")
    output = tmp_path / "complete.xlsx"
    result = run_tool({
        "content_path": "content.md",
        "format_path": "format.md",
        "output_path": "complete.xlsx",
        "overwrite": False,
    }, tmp_path)
    assert result["ok"] is True
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["项目汇总", "评分矩阵"]
    assert workbook["项目汇总"].merged_cells.ranges
    assert len(workbook["评分矩阵"].tables) == 1
    assert len(workbook["评分矩阵"].data_validations.dataValidation) == 0
    assert workbook["项目汇总"]["D3"].value == 12
    assert str(workbook["项目汇总"]["D7"].value).startswith("=SUM")


def test_hidden_format_is_used(tmp_path):
    content = Path(__file__).parent.parent / "examples" / "hidden-format-content.md"
    local_content = tmp_path / "content.md"
    local_content.write_text(content.read_text(encoding="utf-8"), encoding="utf-8")
    output = tmp_path / "hidden.xlsx"
    result = run_tool({
        "content_path": "content.md",
        "output_path": "hidden.xlsx",
    }, tmp_path)
    assert result["ok"] is True
    workbook = load_workbook(output)
    assert workbook["隐藏格式示例"]["A1"].font.bold is True


def test_shape_mismatch_is_rejected(tmp_path):
    content = tmp_path / "content.md"
    content.write_text(CONTENT.read_text(encoding="utf-8"), encoding="utf-8")
    bad_format = tmp_path / "bad-format.md"
    bad_format.write_text("# 项目汇总\n\n| bold |\n|---|\n|  |\n", encoding="utf-8")
    result = run_tool({
        "content_path": "content.md",
        "format_path": str(bad_format),
        "output_path": "bad.xlsx",
    }, tmp_path)
    assert result["ok"] is False
    assert result["error_info"]["code"] == "TABLE_SHAPE_MISMATCH"


def test_bare_unknown_directive_is_rejected(tmp_path):
    content = tmp_path / "content.md"
    content.write_text(CONTENT.read_text(encoding="utf-8"), encoding="utf-8")
    bad_format = tmp_path / "bad-format.md"
    bad_format.write_text("# 项目汇总\n\n| typo-bold |  |  |  |  |  |  |  |\n|---|---|---|---|---|---|---|---|\n" + "|  |  |  |  |  |  |  |  |\n" * 6, encoding="utf-8")
    result = run_tool({
        "content_path": "content.md",
        "format_path": str(bad_format),
        "output_path": "bad.xlsx",
    }, tmp_path)
    assert result["ok"] is False
    assert result["error_info"]["code"] == "INVALID_FORMAT"


def test_chart_and_condition_rule_are_written(tmp_path):
    (tmp_path / "content.md").write_text(
        "# Chart\n\n| Name | A | B |\n|---|---:|---:|\n| X | 10 | 20 |\n| Y | 30 | 40 |\n",
        encoding="utf-8",
    )
    (tmp_path / "format.md").write_text(
        "# Chart\n\n"
        "| chart:bar@A1:C3; chart-title:Demo; cell-is:>20\\|#FFFF00@B2:C3 |  |  |\n"
        "|---|---|---|\n|  |  |  |\n|  |  |  |\n",
        encoding="utf-8",
    )
    result = run_tool({"content_path": "content.md", "format_path": "format.md", "output_path": "chart.xlsx"}, tmp_path)
    assert result["ok"] is True
    workbook = load_workbook(tmp_path / "chart.xlsx")
    assert len(workbook["Chart"]._charts) == 1
    assert len(workbook["Chart"].conditional_formatting) == 1
